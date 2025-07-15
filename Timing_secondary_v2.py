import pandas as pd
import numpy as np
import os
from PyQt5 import QtGui
from collections import defaultdict
import re
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

class TIMING_secondary:
    def __init__(self, path, result_fields, labels):
        self.path = path
        # All items in the path
        all_items = os.listdir(self.path)
        # Filter out only directories
        directories = [item for item in all_items if os.path.isdir(
            os.path.join(self.path, item))]
        # Find directories that start with 'features'
        features_directories = [
            d for d in directories if d.startswith('features')]
        # Append the directory to the path if one is found
        if features_directories:
            filename = os.path.join(
                path, features_directories[0], 'Table_Exp.txt')
            print(f"Full path to the directory: {filename}")
        else:
            print("No directory starts with 'features'")

        # Load data as df
        self.df = pd.read_csv(filename, sep='\t', header=None, index_col=None)
        # Instead of dropping all duplicates, group by columns 0, 1, and 2
        # and only keep the first expected number of rows per well.
        expected_rows = len(labels)
        self.df = self.df.groupby([0, 1, 2], group_keys=False).apply(
            lambda group: group.iloc[:expected_rows])
        print("Extra duplicate rows trimmed to expected rows per well")

        for i in range(1, len(self.df)):
            if not self.df.iloc[i, [0, 1, 2]].equals(self.df.iloc[i - 1, [0, 1, 2]]):
                self.rows_per_well = i
                break

        self.LastRow = len(self.df)
        self.LastColumn = len(self.df.columns)
        # Assuming each "well" is represented by self.rows_per_well rows
        self.Wells = self.LastRow // self.rows_per_well
        # Fix: Calculate the repeat pattern correctly to match the DataFrame's length
        well_repeat_pattern = np.repeat(
            range(1, self.Wells + 1), self.rows_per_well)
        # If the length of the pattern doesn't match the DataFrame, adjust the last well's count
        if len(well_repeat_pattern) < self.LastRow:
            additional_wells = np.array(
                [self.Wells + 1] * (self.LastRow - len(well_repeat_pattern)))
            well_repeat_pattern = np.concatenate(
                (well_repeat_pattern, additional_wells))
        elif len(well_repeat_pattern) > self.LastRow:
            # Trim the excess
            well_repeat_pattern = well_repeat_pattern[:self.LastRow]

        self.df['Well'] = well_repeat_pattern
        self.df.set_index('Well', inplace=True)

        # Adjusting DataFrame to mimic VBA operations:
        # Inserting a column at the beginning for "UID"
        self.df.insert(0, 'UID', np.nan)
        # Since we're adding a header row later, and columns are 0-indexed in pandas,
        # the equivalent of inserting a column at "E" (the 5th position in a 1-indexed system like Excel)
        # is inserting at the 4th index (after adjusting for the new UID column).
        # Placeholder column at original "E" position
        self.df.insert(4, 'Features', np.nan)

        # Preparing headers
        # Assuming the rest of the columns are data columns, with indices starting from 1
        self.headers = ['UID', 'Block', 'X', 'Y', 'Time_Point'] + \
            [f"{i}" for i in range(1, self.df.shape[1] - 4)]

        # Adding a header row at the top and reassigning column names
        self.df.columns = self.headers

        # Eddie edit
        # self.df = self.df.drop(['73'], axis=1)
        # edit end

        # Initializing other variables as per the VBA script
        # (The initialization here is just to declare variables as they appear in your VBA script.
        # The actual calculations or manipulations based on these variables will need specific logic to be implemented)

        # Variables from the VBA script
        SumContT1 = SumContT2 = SumContT3 = SumContT4 = SumContT5 = 0
        SumAliveContT1 = SumAliveContT2 = SumAliveContT3 = SumAliveContT4 = SumAliveContT5 = 0
        SumD1 = SumD2 = SumD3 = SumD4 = SumD5 = 0
        sum_mot_no = count_mot_no = sum_mot_cont = count_mot_cont = 0
        sum_ar_no = count_ar_no = sum_ar_cont = count_ar_cont = 0
        # Initializing kill and dead counters
        Kill_1_1 = Kill_1_0 = Kill_2_2 = Kill_2_1 = Kill_2_0 = Kill_3_3 = Kill_3_2 = Kill_3_1 = Kill_3_0 = 0
        Dead_1_1 = Dead_1_0 = Dead_2_2 = Dead_2_1 = Dead_2_0 = Dead_3_3 = Dead_3_2 = Dead_3_1 = Dead_3_0 = 0
        # Test and frame variables
        Test1 = Test2 = Test3 = TestSum = Frames = Testframes = 0
        self.Frames = self.LastColumn - 5
        self.Testframes = self.Frames * -1 * 500
        print(self.Testframes)
        print("asdfasdf")
        print(self.Frames)
        # Assuming 'self.df' is your DataFrame and 'Wells' is the number of wells in your file
        # Also assuming that 'self.LastColumn' is defined (replace with your actual last column index)
        # EDInt, DInt, Testframes, ContInt should be defined according to your thresholds
        self.InitialLastColumn = len(self.df.columns)
        # self.set_extra_headers(column_headers)
        self.set_labels(labels)
        self.result_fields = result_fields
        self.result_field_inds = {k: v for v,
                                  k in enumerate(self.result_fields)}

    def set_extra_headers(self, column_headers):
        # Now, add new columns with the specified headers
        for header in column_headers:
            # Or use 0 or another placeholder value as default
            self.df[f"{header}"] = None
        self.extra_header_inds = {k: v for v, k in enumerate(column_headers)}

    def set_labels(self, labels):
        self.labels = labels
        self.label_inds = {k: v for v, k in enumerate(labels)}
        if len(labels) != self.rows_per_well:
            print("Rows per well:", self.rows_per_well,
                  "Number of labels provided:", len(labels))
            print("Number of labels invalid--not equal to rows per well")
        for i in range(1, self.Wells + 1):
            start_row = (i - 1) * self.rows_per_well

            # Adjust 4 to your actual column index for labels
            self.df.iloc[start_row:start_row + self.rows_per_well, 4] = labels
        # Fill null values with 0 (or any other appropriate default value)
        self.df.fillna(0, inplace=True)
        self.df['UID'] = self.df['Block'] * 10000 + \
            self.df['X'] * 100 + self.df['Y']

        # print(self.df.head(25))
        # Main Nanowell Processing Loop
        self.LastColumn = len(self.df.columns)

    def display_df(self, rows=None, cols=None):
        if rows is not None and cols is not None:
            # Display specific rows and columns
            print(self.df.iloc[rows, cols])
        elif rows is not None:
            # Display specific rows
            print(self.df.iloc[rows])
        elif cols is not None:
            # Display specific columns
            print(self.df.iloc[:, cols])
        else:
            # Display entire DataFrame
            print(self.df)

    def sum_timepoints(self, label, start_row):
        #print(self.df.iloc[start_row +
        #      self.label_inds[label], 5:(self.LastColumn-1)])
        return self.df.iloc[start_row + self.label_inds[label], 5:(self.LastColumn-1)].sum()

    def calc_alive_cont(self, label, start_row):
        tot = 0
        # print(self.df.iloc[start_row + self.label_inds["Eff_death"], 5])
        # self.df.iloc[start_row + self.label_inds[label], 5]
        for i in range(5, self.LastColumn - 1):
            if self.df.iloc[start_row + self.label_inds["Eff_death"], i] == 0:
                tot += self.df.iloc[start_row + self.label_inds[label], i]
        # See uid 140204
        return tot

        return (np.logical_xor(self.df.iloc[start_row + self.label_inds["Eff_death"], 5:(self.LastColumn - 1)].astype(int), 1)
                & self.df.iloc[start_row + self.label_inds[label], 5:(self.LastColumn - 1)].astype(int)
                ).astype(int).sum()

    def secondary(self, DInt, EDInt, ContInt, TimInt, Pix):
        eff_entry_infos = []
        no_eff_entry_infos = []

        kill_counter = defaultdict(lambda: defaultdict(int))
        death_counter = defaultdict(lambda: defaultdict(int))
        for i in range(1, self.Wells + 1):
            entry_info = [None] * len(self.result_fields)
            start_row = (i - 1) * self.rows_per_well
            print(f"Processing well {i} ")
            Effsum = self.sum_timepoints("Eff_x", start_row)
            Tar1sum = self.sum_timepoints("Tar1x", start_row)
            Tar2sum = self.sum_timepoints("Tar2x", start_row)
            Tar3sum = self.sum_timepoints("Tar3x", start_row)
            Tar4sum = self.sum_timepoints("Tar4x", start_row)
            Tar5sum = self.sum_timepoints("Tar5x", start_row)
            no_targets = sum([int(t_sum > self.Testframes) for t_sum in [
                              Tar1sum, Tar2sum, Tar3sum, Tar4sum, Tar5sum]])

            eff_present = Effsum > self.Testframes

            for comp_int, death_target_int, death_target in \
                [(EDInt, "Eff_death_int", "Eff_death"),
                 (DInt, "Tar1D_Int", "Tar1D"),
                 (DInt, "Tar2D_Int", "Tar2D"),
                 (DInt, "Tar3D_Int", "Tar3D"),
                 (DInt, "Tar4D_Int", "Tar4D"),
                 (DInt, "Tar5D_Int", "Tar5D")]:
                self.df.iloc[start_row + self.label_inds[death_target], 5] = int(
                    (self.df.iloc[start_row +
                     self.label_inds[death_target_int], 5] >= comp_int)
                    and (self.df.iloc[start_row + self.label_inds[death_target_int], 6] >= comp_int)
                )

            for t_sum, t_cont_int, t_cont in \
                [(Tar1sum, "Cont1_Int", "Cont1"),
                 (Tar2sum, "Cont2_Int", "Cont2"),
                 (Tar3sum, "Cont3_Int", "Cont3"),
                 (Tar4sum, "Cont4_Int", "Cont4"),
                 (Tar5sum, "Cont5_Int", "Cont5")]:
                self.df.iloc[start_row + self.label_inds[t_cont], 5] = int(
                    Effsum > self.Testframes
                    and t_sum > self.Testframes
                    and self.df.iloc[start_row + self.label_inds[t_cont_int], 5] >= ContInt
                    and self.df.iloc[start_row + self.label_inds[t_cont_int], 6] >= ContInt)

            # Index 2 to end - 1
            for K in range(6, self.InitialLastColumn - 2):
                for comp_int, death_target_int, death_target in \
                    [(EDInt, "Eff_death_int", "Eff_death"),
                     (DInt, "Tar1D_Int", "Tar1D"),
                     (DInt, "Tar2D_Int", "Tar2D"),
                     (DInt, "Tar3D_Int", "Tar3D"),
                     (DInt, "Tar4D_Int", "Tar4D"),
                     (DInt, "Tar5D_Int", "Tar5D")]:
                    self.df.iloc[start_row + self.label_inds[death_target], K] = int(
                        self.df.iloc[start_row +
                                     self.label_inds[death_target], K - 1] == 1
                        or (
                            (self.df.iloc[start_row +
                             self.label_inds[death_target_int], K] >= comp_int)
                            and (self.df.iloc[start_row + self.label_inds[death_target_int], K + 1] >= comp_int)
                        )
                    )
                for t_sum, t_cont_int, t_cont in \
                    [(Tar1sum, "Cont1_Int", "Cont1"),
                     (Tar2sum, "Cont2_Int", "Cont2"),
                     (Tar3sum, "Cont3_Int", "Cont3"),
                     (Tar4sum, "Cont4_Int", "Cont4"),
                     (Tar5sum, "Cont5_Int", "Cont5")]:
                    self.df.iloc[start_row + self.label_inds[t_cont], K] = int(
                        Effsum > self.Testframes
                        and t_sum > self.Testframes
                        and ((self.df.iloc[start_row + self.label_inds[t_cont_int], K] >= ContInt
                              and (self.df.iloc[start_row + self.label_inds[t_cont_int], K - 1] >= ContInt
                                   or self.df.iloc[start_row + self.label_inds[t_cont_int], K + 1] >= ContInt)
                              )
                             or (self.df.iloc[start_row + self.label_inds[t_cont_int], K - 1] >= ContInt
                                 and self.df.iloc[start_row + self.label_inds[t_cont_int], K + 1] >= ContInt)
                             )
                    )

            for death_target in \
                ["Eff_death",
                 "Tar1D",
                 "Tar2D",
                 "Tar3D",
                 "Tar4D",
                 "Tar5D"]:
                self.df.iloc[start_row + self.label_inds[death_target],
                             self.InitialLastColumn - 2]\
                    = self.df.iloc[start_row + self.label_inds[death_target], self.InitialLastColumn - 3]

            for t_sum, t_cont_int, t_cont in \
                [(Tar1sum, "Cont1_Int", "Cont1"),
                 (Tar2sum, "Cont2_Int", "Cont2"),
                 (Tar3sum, "Cont3_Int", "Cont3"),
                 (Tar4sum, "Cont4_Int", "Cont4"),
                 (Tar5sum, "Cont5_Int", "Cont5")]:
                self.df.iloc[start_row + self.label_inds[t_cont], self.InitialLastColumn - 2] = int(
                    Effsum > self.Testframes
                    and t_sum > self.Testframes
                    and self.df.iloc[start_row + self.label_inds[t_cont_int], self.InitialLastColumn - 2] >= ContInt
                    and self.df.iloc[start_row + self.label_inds[t_cont_int], self.InitialLastColumn - 3] >= ContInt
                )

            # Initialize sums
            SumContT1 = self.sum_timepoints("Cont1", start_row)
            SumContT2 = self.sum_timepoints("Cont2", start_row)
            SumContT3 = self.sum_timepoints("Cont3", start_row)
            SumContT4 = self.sum_timepoints("Cont4", start_row)
            SumContT5 = self.sum_timepoints("Cont5", start_row)
            SumEffD = self.sum_timepoints("Eff_death", start_row)
            SumD1 = self.sum_timepoints("Tar1D", start_row)
            SumD2 = self.sum_timepoints("Tar2D", start_row)
            SumD3 = self.sum_timepoints("Tar3D", start_row)
            SumD4 = self.sum_timepoints("Tar4D", start_row)
            SumD5 = self.sum_timepoints("Tar5D", start_row)
            SumAliveContT1 = self.calc_alive_cont("Cont1", start_row)
            SumAliveContT2 = self.calc_alive_cont("Cont2", start_row)
            SumAliveContT3 = self.calc_alive_cont("Cont3", start_row)
            SumAliveContT4 = self.calc_alive_cont("Cont4", start_row)
            SumAliveContT5 = self.calc_alive_cont("Cont5", start_row)

            # if self.df.iloc[start_row]['UID'] == 10406:
            #     print(Tar1sum)
            #     print(Tar2sum)
            # tSeek and cumulative contact
            # SumAliveContT1
            self.LastColumn = self.InitialLastColumn
            # print("Column at SumAliveConT1: ", LastColumn+16)

            for cur_alive_cont, cur_cont, cont_str, cum_contact, t_seek in \
                [(SumAliveContT1, SumContT1, "Cont1", "Cum contact 1", "tSeek1"),
                 (SumAliveContT2, SumContT2, "Cont2", "Cum contact 2", "tSeek2"),
                 (SumAliveContT3, SumContT3, "Cont3", "Cum contact 3", "tSeek3"),
                 (SumAliveContT4, SumContT4, "Cont4", "Cum contact 4", "tSeek4"),
                 (SumAliveContT5, SumContT5, "Cont5", "Cum contact 5", "tSeek5")]:
                entry_info[self.result_field_inds[cum_contact]
                           ] = cur_alive_cont * TimInt
                t_seek_val = 0
                if cur_cont > 1:
                    if self.df.iloc[start_row + self.label_inds[cont_str], 5] == 1:
                        t_seek_val = 1 / TimInt
                    else:
                        for i in range(5, self.LastColumn - 1):
                            if self.df.iloc[start_row + self.label_inds[cont_str], i] == 1:
                                t_seek_val = i - 5
                                break
                entry_info[self.result_field_inds[t_seek]
                           ] = t_seek_val * TimInt

            for cur_sum, death_marker, death_field in \
                [(SumEffD, "Eff_death", "tEffDeath"),
                 (SumD1, "Tar1D", "Absol tar death 1"),
                 (SumD2, "Tar2D", "Absol tar death 2"),
                 (SumD3, "Tar3D", "Absol tar death 3"),
                 (SumD4, "Tar4D", "Absol tar death 4"),
                 (SumD5, "Tar5D", "Absol tar death 5")]:
                absol_death_val = -1 / TimInt
                #print(death_marker)
                if cur_sum > 0:
                    for i in range(5, self.LastColumn - 1):
                        if self.df.iloc[start_row + self.label_inds[death_marker], i] == 1:
                            absol_death_val = i - 5
                            break
                entry_info[self.result_field_inds[death_field]
                           ] = absol_death_val * TimInt

            # tDeath: Check if absolute time of death > tSeek.
            # When tseek and tDeath are the same, write tDeath as half of TimInt

            for death_field, t_seek, t_death in \
                [("Absol tar death 1", "tSeek1", "tDeath1"),
                 ("Absol tar death 2", "tSeek2", "tDeath2"),
                 ("Absol tar death 3", "tSeek3", "tDeath3"),
                 ("Absol tar death 4", "tSeek4", "tDeath4"),
                 ("Absol tar death 5", "tSeek5", "tDeath5")]:
                tar_death_val = entry_info[self.result_field_inds[death_field]]
                t_seek_val = entry_info[self.result_field_inds[t_seek]]
                t_death_ind = self.result_field_inds[t_death]
                t_death_val = -3
                if tar_death_val > t_seek_val and t_seek_val > 0:
                    t_death_val = tar_death_val - t_seek_val
                elif (tar_death_val == t_seek_val and t_seek_val > 0) or (tar_death_val == 0 and t_seek_val == 1):
                    t_death_val = TimInt / 2
                elif tar_death_val == 0:
                    t_death_val = -2
                elif tar_death_val == -1 and t_seek_val != 0:
                    t_death_val = -1
                entry_info[t_death_ind] = t_death_val

            # for death_field, t_seek, t_death in \
            #     [("Absol tar death 1", "tSeek1", "tDeath1"),
            #      ("Absol tar death 2", "tSeek2", "tDeath2"),
            #      ("Absol tar death 3", "tSeek3", "tDeath3"),
            #      ("Absol tar death 4", "tSeek4", "tDeath4"),
            #      ("Absol tar death 5", "tSeek5", "tDeath5")]:
            #     tar_death_val = entry_info[self.result_field_inds[death_field]]
            #     t_seek_val = entry_info[self.result_field_inds[t_seek]]
            #     t_death_ind = self.result_field_inds[t_death]
            #     t_death_val = -3
            #     if tar_death_val > t_seek_val and t_seek_val > 0:
            #         t_death_val = tar_death_val - t_seek_val
            #     elif (tar_death_val == t_seek_val and t_seek_val > 0) or (tar_death_val == 0 and t_seek_val == 1):
            #         t_death_val = TimInt / 2
            #     elif tar_death_val == 0:
            #         t_death_val = -2
            #     elif tar_death_val == -1 and t_seek_val != 0:
            #         t_death_val = -1
            #     entry_info[t_death_ind] = t_death_val

            ############
            # 'tContact
            for t_death, t_contact, t_seek, t_cont, t_abs_death, sum_cont in \
                [("tDeath1", "tContact1", "tSeek1", "Cont1", "Absol tar death 1", SumContT1),
                 ("tDeath2", "tContact2", "tSeek2",
                  "Cont2", "Absol tar death 2", SumContT2),
                 ("tDeath3", "tContact3", "tSeek3",
                  "Cont3", "Absol tar death 3", SumContT3),
                 ("tDeath4", "tContact4", "tSeek4",
                  "Cont4", "Absol tar death 4", SumContT4),
                 ("tDeath5", "tContact5", "tSeek5", "Cont5", "Absol tar death 5", SumContT5)]:
                t_death_val = entry_info[self.result_field_inds[t_death]]
                t_seek_val = entry_info[self.result_field_inds[t_seek]]
                t_abs_death_val = entry_info[self.result_field_inds[t_abs_death]]
                t_contact_ind = self.result_field_inds[t_contact]
                t_contact_val = -1
                if sum_cont > 1:
                    t_cont_sum = 0
                    for L in range(5 + int((t_seek_val // TimInt)), 6 + int((t_abs_death_val // TimInt))):
                        if self.df.iloc[start_row + self.label_inds[t_cont], L] == 1:
                            t_cont_sum += 1
                            if self.df.iloc[start_row, 0] == 40401 or self.df.iloc[start_row, 0] == 280104:
                                #print(self.df.iloc[start_row, 0])
                                #print(
                                #    self.df.iloc[start_row + self.label_inds[t_cont] - 1, L])
                                #print(
                                #    self.df.iloc[start_row + self.label_inds[t_cont] - 3, L])
                                #print(L)
                                #print((t_cont_sum - 1))
                                print(t_death)
                    if t_death_val > 0:
                        if t_death_val == TimInt / 2:
                            t_contact_val = TimInt / 2
                        else:
                            t_contact_val = (t_cont_sum - 1) * TimInt
                    elif t_abs_death_val == 0 and t_seek_val == 1:
                        t_contact_val = TimInt / 2
                    elif t_death_val == -1:
                        t_cont_sum = 0
                        for L in range(5 + int((t_seek_val // TimInt)), self.LastColumn - 1):
                            if self.df.iloc[start_row + self.label_inds[t_cont], L] == 1:
                                t_cont_sum += 1
                                if self.df.iloc[start_row, 0] == 40401 or self.df.iloc[start_row, 0] == 280104:
                                    #print(self.df.iloc[start_row, 0])
                                    #print(
                                    #    self.df.iloc[start_row + self.label_inds[t_cont] - 1, L])
                                    #print(
                                    #    self.df.iloc[start_row + self.label_inds[t_cont] - 3, L])
                                    #print(L)
                                    #print((t_cont_sum - 1))
                                    print(t_death)
                        t_contact_val = (t_cont_sum - 1) * TimInt
                    else:
                        t_contact_val = -2

                entry_info[t_contact_ind] = t_contact_val

            # def ensure_dataframe_size(df, required_row, required_column):
            #     # Adjust the index to zero-based by subtracting 1 since pandas uses zero-based indexing
            #     required_row_index = required_row - 1
            #     required_column_index = required_column - 1

            #     # Check if we need more rows
            #     while len(df) <= required_row_index:
            #         # Append a new row with None values
            #         df.loc[len(df)] = [None] * len(df.columns)

            #     # Check if we need more columns
            #     while len(df.columns) <= required_column_index:
            #         # Append a new column with None values
            #         df[len(df.columns)] = None

            #     return df

            ##########

            # # Total targets
            # Tar = df.iloc[start_row + 6, LastColumn + 1] + df.iloc[start_row + 14, LastColumn + 1] + df.iloc[start_row +
            #  22, LastColumn + 1] + df.iloc[start_row + 30, LastColumn + 1] + df.iloc[start_row + 38, LastColumn + 1]

            # for j in range(1, 51):
            #     required_row = start_row + j
            #     df = ensure_dataframe_size(df, required_row, LastColumn + 3)

            #     # Now you can safely assign values without encountering IndexError
            #     df.iloc[required_row - 1, LastColumn + 2] = Tar

            #     # Identifying effector
            #     if Effsum > Testframes:
            #         df.iloc[start_row + j - 1, LastColumn] = 1
            #     else:
            #         df.iloc[start_row + j - 1, LastColumn] = 0

            # Effector features
            all_conts = ["Cont1", "Cont2", "Cont3", "Cont4", "Cont5"]
            for proceed_bool, t_abs_death, conts, t_speed, t_ar, d_no_cont, d_cont, ar_no_cont, ar_cont in \
                [[True, "tEffDeath", all_conts, "Eff_Speed", "Eff_AR", "Eff dWell No Cont (um)", "Eff dWell Cont (um)", "Eff AR No Cont", "Eff AR Cont"],
                 [no_targets > 0, "Absol tar death 1", [all_conts[0]], "Tar1_Speed", "Target1_AR",
                     "Tar1 dWell No Cont (um)", "Tar1 dWell Cont (um)", "Tar1 AR No Cont", "Tar1 AR Cont"],
                 [no_targets > 1, "Absol tar death 2", [all_conts[1]], "Tar2_Speed", "Target2_AR",
                     "Tar2 dWell No Cont (um)", "Tar2 dWell Cont (um)", "Tar2 AR No Cont", "Tar2 AR Cont"],
                 [no_targets > 2, "Absol tar death 3", [all_conts[2]], "Tar3_Speed", "Target3_AR",
                     "Tar3 dWell No Cont (um)", "Tar3 dWell Cont (um)", "Tar3 AR No Cont", "Tar3 AR Cont"],
                 [no_targets > 3, "Absol tar death 4", [all_conts[3]], "Tar4_Speed", "Target4_AR",
                     "Tar4 dWell No Cont (um)", "Tar4 dWell Cont (um)", "Tar4 AR No Cont", "Tar4 AR Cont"],
                 [no_targets > 4, "Absol tar death 5", [all_conts[4]], "Tar5_Speed", "Target5_AR", "Tar5 dWell No Cont (um)", "Tar5 dWell Cont (um)", "Tar5 AR No Cont", "Tar5 AR Cont"]]:
                if proceed_bool:
                    sum_mot_no = 0
                    count_mot_no = 0
                    sum_mot_cont = 0
                    count_mot_cont = 0
                    sum_ar_no = 0
                    count_ar_no = 0
                    sum_ar_cont = 0
                    count_ar_cont = 0

                    actual_absol_time = entry_info[self.result_field_inds[t_abs_death]] / TimInt

                    cur_range = range(5, self.LastColumn - 1)
                    if actual_absol_time > 0:
                        cur_range = range(5, int(actual_absol_time) + 5)

                    for K in cur_range:
                        cont_sum = sum(
                            [self.df.iloc[start_row + self.label_inds[cur_cont], K] for cur_cont in conts])
                        t_speed_val = self.df.iloc[start_row +
                                                   self.label_inds[t_speed], K]
                        t_ar_val = self.df.iloc[start_row +
                                                self.label_inds[t_ar], K]
                        if cont_sum == 0:
                            if t_speed_val > 0:
                                sum_mot_no += t_speed_val
                                count_mot_no += 1
                            if t_ar_val > 0:
                                sum_ar_no += t_ar_val
                                count_ar_no += 1
                        else:
                            if t_speed_val > 0:
                                sum_mot_cont += t_speed_val
                                count_mot_cont += 1
                            if t_ar_val > 0:
                                sum_ar_cont += t_ar_val
                                count_ar_cont += 1
                    entry_info[self.result_field_inds[d_no_cont]] = Pix * \
                        sum_mot_no / count_mot_no if count_mot_no > 4 else 0
                    entry_info[self.result_field_inds[d_cont]] = Pix * \
                        sum_mot_cont / count_mot_cont if count_mot_cont > 4 else 0
                    entry_info[self.result_field_inds[ar_no_cont]
                               ] = sum_ar_no / count_ar_no if count_ar_no > 4 else 0
                    entry_info[self.result_field_inds[ar_cont]] = sum_ar_cont / \
                        count_ar_cont if count_ar_cont > 4 else 0

            # # 0E1T and 1E0T analysis
            # entry_info[self.result_field_inds[]]

            # if eff_present:
            #     kills = [int(entry_info[self.result_field_inds[t_death]] > 0 and
            #                  entry_info[self.result_field_inds[t_contact]] > 0) for t_death, t_contact in
            #              [["tDeath1", "tContact1"],
            #               ["tDeath2", "tContact2"],
            #               ["tDeath3", "tContact3"],
            #               ["tDeath4", "tContact4"],
            #               ["tDeath5", "tContact5"],]]
            #     kill_count = sum(kills[:no_targets])

            #     deaths = [int(entry_info[self.result_field_inds[t_death]] != -1 and
            #                   entry_info[self.result_field_inds[t_contact]] != -2) for t_death, t_contact in
            #               [["tDeath1", "tContact1"],
            #               ["tDeath2", "tContact2"],
            #               ["tDeath3", "tContact3"],
            #               ["tDeath4", "tContact4"],
            #               ["tDeath5", "tContact5"],]]
            #     death_count = sum(deaths[:no_targets])

            #     kill_counter[no_targets][kill_count] += 1
            #     death_counter[no_targets][death_count] += 1

            entry_info[self.result_field_inds["UID"]
                       ] = self.df.iloc[start_row]['UID']
            entry_info[self.result_field_inds["No of targets"]] = no_targets

            if Effsum > self.Testframes:
                eff_entry_infos.append(entry_info)
                continue
            no_eff_entry_infos.append(entry_info)
        # print(all_entry_infos)

        home_directory = os.path.join(self.path, 'secondary_results')
        os.makedirs(home_directory, exist_ok=True)
        indiv_files_directory = os.path.join(
            home_directory, 'individual_files')
        os.makedirs(indiv_files_directory, exist_ok=True)
        # self.df.to_csv(os.path.join(home_directory, "output.csv"),
        #                sep='\t', index=False)
        filename_1E_0T = os.path.join(indiv_files_directory, 'Table_1E0T.csv')
        filename_0E_1T = os.path.join(indiv_files_directory, 'Table_0E1T.csv')
        filename_1E_1T = os.path.join(indiv_files_directory, 'Table_1E1T.csv')
        filename_1E_2T = os.path.join(indiv_files_directory, 'Table_1E2T.csv')
        filename_1E_3T = os.path.join(indiv_files_directory, 'Table_1E3T.csv')
        filename_1E_PT = os.path.join(indiv_files_directory, 'Table_1E1+T.csv')
        #print(eff_entry_infos)

        df_0E_1T = self.format_0E(pd.DataFrame([x for x in no_eff_entry_infos if x[self.result_field_inds["No of targets"]] == 1],
                                               columns=self.result_fields))
        df_1E_0T = self.format_0T(pd.DataFrame([x for x in eff_entry_infos if x[self.result_field_inds["No of targets"]] == 0],
                                               columns=self.result_fields))
        df_1E_1T = self.format_1E(pd.DataFrame([x for x in eff_entry_infos if x[self.result_field_inds["No of targets"]] == 1],
                                               columns=self.result_fields), 1)
        df_1E_2T = self.format_1E(pd.DataFrame([x for x in eff_entry_infos if x[self.result_field_inds["No of targets"]] == 2],
                                               columns=self.result_fields), 2)
        df_1E_3T = self.format_1E(pd.DataFrame([x for x in eff_entry_infos if x[self.result_field_inds["No of targets"]] == 3],
                                               columns=self.result_fields), 3)
        df_1E_PT = pd.DataFrame([x for x in eff_entry_infos if x[self.result_field_inds["No of targets"]] > 0],
                                columns=self.result_fields)

        # 0E1T and 1E0T analysis
        # E | T | Valid | Contact | Death
        # df_0E_1T_tot = df_0E_1T.shape[0]
        df_0E_1T_death = df_0E_1T[df_0E_1T['TarDeath'] >= 0]
        df_0E_1T_no_death = df_0E_1T[df_0E_1T['TarDeath'] < 0]
        df_0E_1T_analysis = {
            'tot': df_0E_1T,
            'death': df_0E_1T_death,
            'no_death': df_0E_1T_no_death
        }

        # df_1E_0T_tot = df_1E_0T.shape[0]
        df_1E_0T_eff_death = df_1E_0T[df_1E_0T['tEffDeath'] >= 0]
        df_1E_0T_no_eff_death = df_1E_0T[df_1E_0T['tEffDeath'] < 0]
        df_1E_0T_analysis = {
            'tot': df_1E_0T,
            'death': df_1E_0T_eff_death,
            'no_death': df_1E_0T_no_eff_death
        }

        # df_1E_1T_tot = df_1E_1T.shape[0]
        df_1E_1T_valid = df_1E_1T[df_1E_1T['tDeath1'] != -2]
        df_1E_1T_valid_contact = df_1E_1T[
            (df_1E_1T['tDeath1'] != -2) & (df_1E_1T['tContact1'] >= 0)]
        df_1E_1T_valid_no_contact = df_1E_1T[
            (df_1E_1T['tDeath1'] != -2) & (df_1E_1T['tContact1'] < 0)]
        df_1E_1T_valid_contact_kill = df_1E_1T[
            (df_1E_1T['tDeath1'] != -2) & (df_1E_1T['tContact1'] >= 0) & (df_1E_1T['tDeath1'] >= 0)]
        df_1E_1T_valid_contact_no_kill = df_1E_1T[
            (df_1E_1T['tDeath1'] != -2) & (df_1E_1T['tContact1'] >= 0) & (df_1E_1T['tDeath1'] < 0)]
        df_1E_1T_valid_no_contact_death = df_1E_1T[
            (df_1E_1T['tDeath1'] != -2) & (df_1E_1T['tContact1'] < 0) & (df_1E_1T['Absol tar death 1'] >= 0)]
        df_1E_1T_valid_no_contact_no_death = df_1E_1T[
            (df_1E_1T['tDeath1'] != -2) & (df_1E_1T['tContact1'] < 0) & (df_1E_1T['Absol tar death 1'] < 0)]
        df_1E_1T_analysis = {
            'valid': df_1E_1T_valid,
            'valid_contact': df_1E_1T_valid_contact,
            'valid_no_contact': df_1E_1T_valid_no_contact,
            'valid_contact_kill': df_1E_1T_valid_contact_kill,
            'valid_contact_no_kill': df_1E_1T_valid_contact_no_kill,
            'valid_no_contact_death': df_1E_1T_valid_no_contact_death,
            'valid_no_contact_no_death': df_1E_1T_valid_no_contact_no_death,
            'tot': df_1E_1T,
        }

        # df_1E_2T_valid = df_1E_2T[
        #     (df_1E_2T['tDeath1'] != -2) | (df_1E_2T['tDeath2'] != -2)]
        # df_1E_2T_valid_contact = df_1E_2T[
        #     ((df_1E_2T['tDeath1'] != -2) & (df_1E_2T['tContact1'] >= 0)) | ((df_1E_2T['tDeath2'] != -2) & (df_1E_2T['tContact2'] >= 0))]
        # df_1E_2T_valid_no_contact = df_1E_2T[
        #     ((df_1E_2T['tDeath1'] != -2) & (df_1E_2T['tContact1'] < 0)) & ((df_1E_2T['tDeath2'] != -2) & (df_1E_2T['tContact2'] < 0))]
        # df_1E_2T_valid_contact_kill = df_1E_2T[
        #     ((df_1E_2T['tDeath1'] != -2) & (df_1E_2T['tContact1'] >= 0) & (df_1E_2T['tDeath1'] >= 0)) | ((df_1E_2T['tDeath2'] != -2) & (df_1E_2T['tContact2'] >= 0) & (df_1E_2T['tDeath2'] >= 0))]
        # df_1E_2T_valid_contact_no_kill = df_1E_2T[
        #     ((df_1E_2T['tDeath1'] != -2) & (df_1E_2T['tContact1'] >= 0) & (df_1E_2T['tDeath1'] < 0)) & ((df_1E_2T['tDeath2'] != -2) & (df_1E_2T['tContact2'] >= 0) & (df_1E_2T['tDeath2'] < 0))]
        # df_1E_2T_valid_no_contact_death = df_1E_2T[
        #     ((df_1E_2T['tDeath1'] != -2) & (df_1E_2T['tContact1'] < 0) & (df_1E_2T['Absol tar death 1'] >= 0)) | ((df_1E_2T['tDeath2'] != -2) & (df_1E_2T['tContact2'] < 0) & (df_1E_2T['Absol tar death 2'] >= 0))]
        # df_1E_2T_valid_no_contact_no_death = df_1E_2T[
        #     ((df_1E_2T['tDeath1'] != -2) & (df_1E_2T['tContact1'] < 0) & (df_1E_2T['Absol tar death 1'] < 0)) & ((df_1E_2T['tDeath2'] != -2) & (df_1E_2T['tContact2'] < 0) & (df_1E_2T['Absol tar death 2'] < 0))]
        # df_1E_2T_analysis = {
        #     'valid': df_1E_2T_valid,
        #     'valid_contact': df_1E_2T_valid_contact,
        #     'valid_no_contact': df_1E_2T_valid_no_contact,
        #     'valid_contact_kill': df_1E_2T_valid_contact_kill,
        #     'valid_contact_no_kill': df_1E_2T_valid_contact_no_kill,
        #     'valid_no_contact_death': df_1E_2T_valid_no_contact_death,
        #     'valid_no_contact_no_death': df_1E_2T_valid_no_contact_no_death,
        #     'tot': df_1E_2T,
        # }

        # df_1E_3T_valid = df_1E_3T[
        #     (df_1E_3T['tDeath1'] != -2) | (df_1E_3T['tDeath2'] != -2) | (df_1E_3T['tDeath2'] != -2)]
        # df_1E_3T_valid_contact = df_1E_3T[
        #     ((df_1E_3T['tDeath1'] != -2) & (df_1E_3T['tContact1'] >= 0)) | ((df_1E_3T['tDeath2'] != -2) & (df_1E_3T['tContact2'] >= 0)) | ((df_1E_3T['tDeath3'] != -2) & (df_1E_3T['tContact3'] >= 0))]
        # df_1E_3T_valid_no_contact = df_1E_3T[
        #     ((df_1E_3T['tDeath1'] != -2) & (df_1E_3T['tContact1'] < 0)) & ((df_1E_3T['tDeath2'] != -2) & (df_1E_3T['tContact2'] < 0)) & ((df_1E_3T['tDeath3'] != -2) & (df_1E_3T['tContact3'] < 0))]
        # df_1E_3T_valid_contact_kill = df_1E_3T[
        #     ((df_1E_3T['tDeath1'] != -2) & (df_1E_3T['tContact1'] >= 0) & (df_1E_3T['tDeath1'] >= 0)) | ((df_1E_3T['tDeath2'] != -2) & (df_1E_3T['tContact2'] >= 0) & (df_1E_3T['tDeath2'] >= 0)) | ((df_1E_3T['tDeath3'] != -2) & (df_1E_3T['tContact3'] >= 0) & (df_1E_3T['tDeath3'] >= 0))]
        # df_1E_3T_valid_contact_no_kill = df_1E_3T[
        #     ((df_1E_3T['tDeath1'] != -2) & (df_1E_3T['tContact1'] >= 0) & (df_1E_3T['tDeath1'] < 0)) & ((df_1E_3T['tDeath2'] != -2) & (df_1E_3T['tContact2'] >= 0) & (df_1E_3T['tDeath2'] < 0)) & ((df_1E_3T['tDeath3'] != -2) & (df_1E_3T['tContact3'] >= 0) & (df_1E_3T['tDeath3'] < 0))]
        # df_1E_3T_valid_no_contact_death = df_1E_3T[
        #     ((df_1E_3T['tDeath1'] != -2) & (df_1E_3T['tContact1'] < 0) & (df_1E_3T['Absol tar death 1'] >= 0)) | ((df_1E_3T['tDeath2'] != -2) & (df_1E_3T['tContact2'] < 0) & (df_1E_3T['Absol tar death 2'] >= 0)) | ((df_1E_3T['tDeath3'] != -2) & (df_1E_3T['tContact3'] < 0) & (df_1E_3T['Absol tar death 3'] >= 0))]
        # df_1E_3T_valid_no_contact_no_death = df_1E_3T[
        #     ((df_1E_3T['tDeath1'] != -2) & (df_1E_3T['tContact1'] < 0) & (df_1E_3T['Absol tar death 1'] < 0)) & ((df_1E_3T['tDeath2'] != -2) & (df_1E_3T['tContact2'] < 0) & (df_1E_3T['Absol tar death 2'] < 0)) & ((df_1E_3T['tDeath3'] != -2) & (df_1E_3T['tContact3'] < 0) & (df_1E_3T['Absol tar death 3'] < 0))]
        # df_1E_3T_analysis = {
        #     'valid': df_1E_3T_valid,
        #     'valid_contact': df_1E_3T_valid_contact,
        #     'valid_no_contact': df_1E_3T_valid_no_contact,
        #     'valid_contact_kill': df_1E_3T_valid_contact_kill,
        #     'valid_contact_no_kill': df_1E_3T_valid_contact_no_kill,
        #     'valid_no_contact_death': df_1E_3T_valid_no_contact_death,
        #     'valid_no_contact_no_death': df_1E_3T_valid_no_contact_no_death,
        #     'tot': df_1E_3T,
        # }

        # df_1E_2T_valid = df_1E_2T[
        #     (df_1E_2T['tDeath1'] != -2) & (df_1E_2T['tDeath2'] != -2)]
        # df_1E_2T_valid_contact = df_1E_2T_valid[
        #     (df_1E_2T_valid['tContact1'] >= 0) | (df_1E_2T_valid['tContact2'] >= 0)]
        # df_1E_2T_valid_no_contact = df_1E_2T_valid[
        #     (df_1E_2T_valid['tContact1'] < 0) & (df_1E_2T_valid['tContact2'] < 0)]
        # df_1E_2T_valid_contact_kill = df_1E_2T_valid[
        #     ((df_1E_2T_valid['tContact1'] >= 0) & (df_1E_2T_valid['tDeath1'] >= 0)) | ((df_1E_2T_valid['tContact2'] >= 0) & (df_1E_2T_valid['tDeath2'] >= 0))]
        # df_1E_2T_valid_contact_no_kill = df_1E_2T_valid[
        #     ((df_1E_2T_valid['tContact1'] >= 0) & (df_1E_2T_valid['tDeath1'] < 0)) & ((df_1E_2T_valid['tContact2'] >= 0) & (df_1E_2T_valid['tDeath2'] < 0))]
        # df_1E_2T_valid_no_contact_death = df_1E_2T_valid[
        #     ((df_1E_2T_valid['tContact1'] < 0) & (df_1E_2T_valid['Absol tar death 1'] >= 0)) | ((df_1E_2T_valid['tContact2'] < 0) & (df_1E_2T_valid['Absol tar death 2'] >= 0))]
        # df_1E_2T_valid_no_contact_no_death = df_1E_2T_valid[
        #     ((df_1E_2T_valid['tContact1'] < 0) & (df_1E_2T_valid['Absol tar death 1'] < 0)) & ((df_1E_2T_valid['tContact2'] < 0) & (df_1E_2T_valid['Absol tar death 2'] < 0))]
        # df_1E_2T_analysis = {
        #     'valid': df_1E_2T_valid,
        #     'valid_contact': df_1E_2T_valid_contact,
        #     'valid_no_contact': df_1E_2T_valid_no_contact,
        #     'valid_contact_kill': df_1E_2T_valid_contact_kill,
        #     'valid_contact_no_kill': df_1E_2T_valid_contact_no_kill,
        #     'valid_no_contact_death': df_1E_2T_valid_no_contact_death,
        #     'valid_no_contact_no_death': df_1E_2T_valid_no_contact_no_death,
        #     'tot': df_1E_2T,
        # }
        df_1E_2T_valid = df_1E_2T[
            (df_1E_2T['tDeath1'] != -2) & (df_1E_2T['tDeath2'] != -2)]
        df_1E_2T_valid_1_contact = df_1E_2T_valid[
            (df_1E_2T_valid['tContact1'] >= 0) & (df_1E_2T_valid['tContact2'] < 0) | (df_1E_2T_valid['tContact1'] < 0) & (df_1E_2T_valid['tContact2'] >= 0)]
        df_1E_2T_valid_2_contact = df_1E_2T_valid[
            (df_1E_2T_valid['tContact1'] >= 0) & (df_1E_2T_valid['tContact2'] >= 0)]
        df_1E_2T_valid_no_contact = df_1E_2T_valid[
            (df_1E_2T_valid['tContact1'] < 0) & (df_1E_2T_valid['tContact2'] < 0)]
        df_1E_2T_valid_2_contact_2_kill = df_1E_2T_valid[
            ((df_1E_2T_valid['tContact1'] >= 0) & (df_1E_2T_valid['tDeath1'] >= 0)) & ((df_1E_2T_valid['tContact2'] >= 0) & (df_1E_2T_valid['tDeath2'] >= 0))]
        df_1E_2T_valid_2_contact_1_kill = df_1E_2T_valid[
            (((df_1E_2T_valid['tContact1'] >= 0) & (df_1E_2T_valid['tDeath1'] >= 0)) & ((df_1E_2T_valid['tContact2'] >= 0) & (df_1E_2T_valid['tDeath2'] < 0))) |
            (((df_1E_2T_valid['tContact1'] >= 0) & (df_1E_2T_valid['tDeath1'] < 0)) & ((df_1E_2T_valid['tContact2'] >= 0) & (df_1E_2T_valid['tDeath2'] >= 0)))]
        df_1E_2T_valid_2_contact_no_kill = df_1E_2T_valid[
            ((df_1E_2T_valid['tContact1'] >= 0) & (df_1E_2T_valid['tDeath1'] < 0)) & ((df_1E_2T_valid['tContact2'] >= 0) & (df_1E_2T_valid['tDeath2'] < 0))]
        df_1E_2T_valid_1_contact_1_kill = df_1E_2T_valid[
            (((df_1E_2T_valid['tContact1'] >= 0) & (df_1E_2T_valid['tDeath1'] >= 0)) & (df_1E_2T_valid['tContact2'] < 0)) |
            (((df_1E_2T_valid['tContact2'] >= 0) & (df_1E_2T_valid['tDeath2'] >= 0)) & (df_1E_2T_valid['tContact1'] < 0))]
        df_1E_2T_valid_1_contact_no_kill = df_1E_2T_valid[
            (((df_1E_2T_valid['tContact1'] >= 0) & (df_1E_2T_valid['tDeath1'] < 0)) & (df_1E_2T_valid['tContact2'] < 0)) |
            (((df_1E_2T_valid['tContact2'] >= 0) & (df_1E_2T_valid['tDeath2'] < 0)) & (df_1E_2T_valid['tContact1'] < 0))]
        df_1E_2T_valid_no_contact_2_death = df_1E_2T_valid[
            ((df_1E_2T_valid['tContact1'] < 0) & (df_1E_2T_valid['Absol tar death 1'] >= 0)) & ((df_1E_2T_valid['tContact2'] < 0) & (df_1E_2T_valid['Absol tar death 2'] >= 0))]
        df_1E_2T_valid_no_contact_1_death = df_1E_2T_valid[
            ((df_1E_2T_valid['tContact1'] < 0) & (df_1E_2T_valid['Absol tar death 1'] >= 0)) & ((df_1E_2T_valid['tContact2'] < 0) & (df_1E_2T_valid['Absol tar death 2'] < 0)) |
            ((df_1E_2T_valid['tContact1'] < 0) & (df_1E_2T_valid['Absol tar death 1'] < 0)) & ((df_1E_2T_valid['tContact2'] < 0) & (df_1E_2T_valid['Absol tar death 2'] >= 0))]
        df_1E_2T_valid_no_contact_no_death = df_1E_2T_valid[
            ((df_1E_2T_valid['tContact1'] < 0) & (df_1E_2T_valid['Absol tar death 1'] < 0)) & ((df_1E_2T_valid['tContact2'] < 0) & (df_1E_2T_valid['Absol tar death 2'] < 0))]
        df_1E_2T_analysis = {
            'valid': df_1E_2T_valid,
            'valid_2_contact': df_1E_2T_valid_2_contact,
            'valid_1_contact': df_1E_2T_valid_1_contact,
            'valid_no_contact': df_1E_2T_valid_no_contact,
            'valid_2_contact_2_kill': df_1E_2T_valid_2_contact_2_kill,
            'valid_2_contact_1_kill': df_1E_2T_valid_2_contact_1_kill,
            'valid_2_contact_no_kill': df_1E_2T_valid_2_contact_no_kill,
            'valid_1_contact_1_kill': df_1E_2T_valid_1_contact_1_kill,
            'valid_1_contact_no_kill': df_1E_2T_valid_1_contact_no_kill,
            'valid_no_contact_2_death': df_1E_2T_valid_no_contact_2_death,
            'valid_no_contact_1_death': df_1E_2T_valid_no_contact_1_death,
            'valid_no_contact_no_death': df_1E_2T_valid_no_contact_no_death,
            'tot': df_1E_2T,
        }


        df_1E_3T_valid = df_1E_3T[
            (df_1E_3T['tDeath1'] != -2) & (df_1E_3T['tDeath2'] != -2) & (df_1E_3T['tDeath2'] != -2)]
        df_1E_3T_valid_3_contact = df_1E_3T_valid[
            (df_1E_3T_valid['tContact1'] >= 0) & (df_1E_3T_valid['tContact2'] >= 0) & (df_1E_3T_valid['tContact3'] >= 0)]
        df_1E_3T_valid_2_contact = df_1E_3T_valid[
            (df_1E_3T_valid['tContact1'] < 0) & (df_1E_3T_valid['tContact2'] >= 0) & (df_1E_3T_valid['tContact3'] >= 0) |
            (df_1E_3T_valid['tContact1'] >= 0) & (df_1E_3T_valid['tContact2'] < 0) & (df_1E_3T_valid['tContact3'] >= 0) |
            (df_1E_3T_valid['tContact1'] >= 0) & (df_1E_3T_valid['tContact2'] >= 0) & (df_1E_3T_valid['tContact3'] < 0)]
        df_1E_3T_valid_1_contact = df_1E_3T_valid[
            (df_1E_3T_valid['tContact1'] < 0) & (df_1E_3T_valid['tContact2'] < 0) & (df_1E_3T_valid['tContact3'] >= 0) |
            (df_1E_3T_valid['tContact1'] >= 0) & (df_1E_3T_valid['tContact2'] < 0) & (df_1E_3T_valid['tContact3'] < 0) |
            (df_1E_3T_valid['tContact1'] < 0) & (df_1E_3T_valid['tContact2'] >= 0) & (df_1E_3T_valid['tContact3'] < 0)]
        df_1E_3T_valid_no_contact = df_1E_3T_valid[
            (df_1E_3T_valid['tContact1'] < 0) & (df_1E_3T_valid['tContact2'] < 0) & (df_1E_3T_valid['tContact3'] < 0)]
        df_1E_3T_valid_p_contact_3_kill = df_1E_3T_valid[
            ((df_1E_3T_valid["tContact1"] >= 0) | (df_1E_3T_valid["tContact2"] >= 0) | (df_1E_3T_valid["tContact2"] >= 0)) &
            ((df_1E_3T_valid["tDeath1"] >= 0) & (df_1E_3T_valid["tDeath2"] >= 0) & (df_1E_3T_valid["tDeath3"] >= 0))]
        df_1E_3T_valid_p_contact_2_kill = df_1E_3T_valid[
            ((df_1E_3T_valid["tContact1"] >= 0) | (df_1E_3T_valid["tContact2"] >= 0) | (df_1E_3T_valid["tContact2"] >= 0)) &
            (((df_1E_3T_valid["tDeath1"] < 0) & (df_1E_3T_valid["tDeath2"] >= 0) & (df_1E_3T_valid["tDeath3"] >= 0)) |
             ((df_1E_3T_valid["tDeath1"] >= 0) & (df_1E_3T_valid["tDeath2"] < 0) & (df_1E_3T_valid["tDeath3"] >= 0)) |
             ((df_1E_3T_valid["tDeath1"] >= 0) & (df_1E_3T_valid["tDeath2"] >= 0) & (df_1E_3T_valid["tDeath3"] < 0)))]
        df_1E_3T_valid_p_contact_1_kill = df_1E_3T_valid[
            ((df_1E_3T_valid["tContact1"] >= 0) | (df_1E_3T_valid["tContact2"] >= 0) | (df_1E_3T_valid["tContact2"] >= 0)) &
            (((df_1E_3T_valid["tDeath1"] < 0) & (df_1E_3T_valid["tDeath2"] < 0) & (df_1E_3T_valid["tDeath3"] >= 0)) |
             ((df_1E_3T_valid["tDeath1"] >= 0) & (df_1E_3T_valid["tDeath2"] < 0) & (df_1E_3T_valid["tDeath3"] < 0)) |
             ((df_1E_3T_valid["tDeath1"] < 0) & (df_1E_3T_valid["tDeath2"] >= 0) & (df_1E_3T_valid["tDeath3"] < 0)))]
        df_1E_3T_valid_p_contact_no_kill = df_1E_3T_valid[
            ((df_1E_3T_valid['tContact1'] >= 0) & (df_1E_3T_valid['tDeath1'] < 0)) & ((df_1E_3T_valid['tContact2'] >= 0) & (df_1E_3T_valid['tDeath2'] < 0)) & ((df_1E_3T_valid['tContact3'] >= 0) & (df_1E_3T_valid['tDeath3'] < 0))]
        df_1E_3T_valid_no_contact_3_death = df_1E_3T_valid[
            ((df_1E_3T_valid['tContact1'] < 0) & (df_1E_3T_valid['Absol tar death 1'] >= 0)) & ((df_1E_3T_valid['tContact2'] < 0) & (df_1E_3T_valid['Absol tar death 2'] >= 0)) & ((df_1E_3T_valid['tContact3'] < 0) & (df_1E_3T_valid['Absol tar death 3'] >= 0))]
        df_1E_3T_valid_no_contact_2_death = df_1E_3T_valid[
            ((df_1E_3T_valid['tContact1'] < 0) & (df_1E_3T_valid['Absol tar death 1'] < 0)) & ((df_1E_3T_valid['tContact2'] < 0) & (df_1E_3T_valid['Absol tar death 2'] >= 0)) & ((df_1E_3T_valid['tContact3'] < 0) & (df_1E_3T_valid['Absol tar death 3'] >= 0)) |
            ((df_1E_3T_valid['tContact1'] < 0) & (df_1E_3T_valid['Absol tar death 1'] >= 0)) & ((df_1E_3T_valid['tContact2'] < 0) & (df_1E_3T_valid['Absol tar death 2'] < 0)) & ((df_1E_3T_valid['tContact3'] < 0) & (df_1E_3T_valid['Absol tar death 3'] >= 0)) |
            ((df_1E_3T_valid['tContact1'] < 0) & (df_1E_3T_valid['Absol tar death 1'] >= 0)) & ((df_1E_3T_valid['tContact2'] < 0) & (df_1E_3T_valid['Absol tar death 2'] >= 0)) & ((df_1E_3T_valid['tContact3'] < 0) & (df_1E_3T_valid['Absol tar death 3'] < 0))]
        df_1E_3T_valid_no_contact_1_death = df_1E_3T_valid[
            ((df_1E_3T_valid['tContact1'] < 0) & (df_1E_3T_valid['Absol tar death 1'] < 0)) & ((df_1E_3T_valid['tContact2'] < 0) & (df_1E_3T_valid['Absol tar death 2'] < 0)) & ((df_1E_3T_valid['tContact3'] < 0) & (df_1E_3T_valid['Absol tar death 3'] >= 0)) |
            ((df_1E_3T_valid['tContact1'] < 0) & (df_1E_3T_valid['Absol tar death 1'] >= 0)) & ((df_1E_3T_valid['tContact2'] < 0) & (df_1E_3T_valid['Absol tar death 2'] < 0)) & ((df_1E_3T_valid['tContact3'] < 0) & (df_1E_3T_valid['Absol tar death 3'] < 0)) |
            ((df_1E_3T_valid['tContact1'] < 0) & (df_1E_3T_valid['Absol tar death 1'] < 0)) & ((df_1E_3T_valid['tContact2'] < 0) & (df_1E_3T_valid['Absol tar death 2'] >= 0)) & ((df_1E_3T_valid['tContact3'] < 0) & (df_1E_3T_valid['Absol tar death 3'] < 0))]
        df_1E_3T_valid_no_contact_no_death = df_1E_3T_valid[
            ((df_1E_3T_valid['tContact1'] < 0) & (df_1E_3T_valid['Absol tar death 1'] < 0)) & ((df_1E_3T_valid['tContact2'] < 0) & (df_1E_3T_valid['Absol tar death 2'] < 0)) & ((df_1E_3T_valid['tContact3'] < 0) & (df_1E_3T_valid['Absol tar death 3'] < 0))]
        df_1E_3T_analysis = {
            'valid': df_1E_3T_valid,
            'valid_3_contact': df_1E_3T_valid_3_contact,
            'valid_2_contact': df_1E_3T_valid_2_contact,
            'valid_1_contact': df_1E_3T_valid_1_contact,
            'valid_no_contact': df_1E_3T_valid_no_contact,
            'valid_p_contact_3_kill': df_1E_3T_valid_p_contact_3_kill,
            'valid_p_contact_2_kill': df_1E_3T_valid_p_contact_2_kill,
            'valid_p_contact_1_kill': df_1E_3T_valid_p_contact_1_kill,
            'valid_p_contact_no_kill': df_1E_3T_valid_p_contact_no_kill,
            'valid_no_contact_3_death': df_1E_3T_valid_no_contact_3_death,
            'valid_no_contact_2_death': df_1E_3T_valid_no_contact_2_death,
            'valid_no_contact_1_death': df_1E_3T_valid_no_contact_1_death,
            'valid_no_contact_no_death': df_1E_3T_valid_no_contact_no_death,
            'tot': df_1E_3T,
        }

        # 0E1T organizing
        df_summary = pd.DataFrame(index=range(60), columns=range(60))
        df_summary.iat[1, 1] = '0E 1T'
        # Use biological logic for 0E 1T
        total_0e1t = df_0E_1T.shape[0]
        invalid_0e1t = ((df_0E_1T['TarDeath'] == 0) | (df_0E_1T['TarDeath'] == 5)).sum()
        valid_0e1t = total_0e1t - invalid_0e1t
        alive_0e1t = (df_0E_1T['TarDeath'] == -1).sum()
        dead_0e1t = valid_0e1t - alive_0e1t
        df_summary.iat[2, 1] = 'tot'
        df_summary.iat[2, 2] = total_0e1t
        df_summary.iat[3, 1] = 'valid'
        df_summary.iat[3, 2] = valid_0e1t
        df_summary.iat[4, 1] = 'invalid'
        df_summary.iat[4, 2] = invalid_0e1t
        df_summary.iat[5, 1] = 'alive'
        df_summary.iat[5, 2] = alive_0e1t
        df_summary.iat[6, 1] = 'dead'
        df_summary.iat[6, 2] = dead_0e1t
        df_summary.iat[7, 1] = "% death"
        if total_0e1t != 0:
            df_summary.iat[7, 2] = dead_0e1t / total_0e1t

        # 1E0T organizing
        df_summary.iat[9, 1] = '1E 0T'
        # Use biological logic for 1E 0T
        total_1e0t = df_1E_0T.shape[0]
        invalid_1e0t = ((df_1E_0T['tEffDeath'] == 0) | (df_1E_0T['tEffDeath'] == 5)).sum()
        valid_1e0t = total_1e0t - invalid_1e0t
        alive_1e0t = (df_1E_0T['tEffDeath'] == -1).sum()
        dead_1e0t = valid_1e0t - alive_1e0t
        df_summary.iat[10, 1] = 'tot'
        df_summary.iat[10, 2] = total_1e0t
        df_summary.iat[11, 1] = 'valid'
        df_summary.iat[11, 2] = valid_1e0t
        df_summary.iat[12, 1] = 'invalid'
        df_summary.iat[12, 2] = invalid_1e0t
        df_summary.iat[13, 1] = 'alive'
        df_summary.iat[13, 2] = alive_1e0t
        df_summary.iat[14, 1] = 'dead'
        df_summary.iat[14, 2] = dead_1e0t
        df_summary.iat[15, 1] = "% death"
        if total_1e0t != 0:
            df_summary.iat[15, 2] = dead_1e0t / total_1e0t

        # 1E1T organizing
        df_summary.iat[17, 1] = '1E 1T'
        for i, k in enumerate(['tot', 'valid']):
            df_summary.iat[18 + i, 1] = k
            df_summary.iat[18 + i, 2] = df_1E_1T_analysis[k].shape[0]

        for i, k in enumerate(['valid_contact', 'valid_contact_kill', 'valid_contact_no_kill']):
            df_summary.iat[18 + i, 4] = k
            df_summary.iat[18 + i, 5] = df_1E_1T_analysis[k].shape[0]
        df_summary.iat[21, 4] = "% kill"
        if df_1E_1T_analysis['valid_contact'].shape[0] != 0:
            df_summary.iat[21, 5] = df_1E_1T_analysis['valid_contact_kill'].shape[0] / \
                df_1E_1T_analysis['valid_contact'].shape[0]

        for i, k in enumerate(['valid_no_contact', 'valid_no_contact_death', 'valid_no_contact_no_death']):
            df_summary.iat[18 + i, 7] = k
            df_summary.iat[18 + i, 8] = df_1E_1T_analysis[k].shape[0]
        df_summary.iat[21, 7] = "% death"
        if df_1E_1T_analysis['valid_no_contact'].shape[0] != 0:
            df_summary.iat[21, 8] = df_1E_1T_analysis['valid_no_contact_death'].shape[0] / \
                df_1E_1T_analysis['valid_no_contact'].shape[0]

        # f_1E_2T_analysis = {
        #     'valid': df_1E_2T_valid,
        #     'valid_2_contact': df_1E_2T_valid_2_contact,
        #     'valid_1_contact': df_1E_2T_valid_1_contact,
        #     'valid_no_contact': df_1E_2T_valid_no_contact,
        #     'valid_2_contact_2_kill': df_1E_2T_valid_2_contact_2_kill,
        #     'valid_2_contact_1_kill': df_1E_2T_valid_2_contact_1_kill,
        #     'valid_2_contact_no_kill': df_1E_2T_valid_2_contact_no_kill,
        #     'valid_1_contact_1_kill': df_1E_2T_valid_1_contact_1_kill,
        #     'valid_1_contact_no_kill': df_1E_2T_valid_1_contact_no_kill,
        #     'valid_no_contact_2_death': df_1E_2T_valid_no_contact_2_death,
        #     'valid_no_contact_1_death': df_1E_2T_valid_no_contact_1_death,
        #     'valid_no_contact_no_death': df_1E_2T_valid_no_contact_no_death,
        #     'tot': df_1E_2T,
        # }
        # 1E2T organizing
        df_summary.iat[23, 1] = '1E 2T'
        for i, k in enumerate(['tot', 'valid']):
            df_summary.iat[24 + i, 1] = k
            df_summary.iat[24 + i, 2] = df_1E_2T_analysis[k].shape[0]

        for i, k in enumerate(['valid_2_contact', 'valid_2_contact_2_kill', 'valid_2_contact_1_kill', 'valid_2_contact_no_kill']):
            df_summary.iat[24 + i, 4] = k
            df_summary.iat[24 + i, 5] = df_1E_2T_analysis[k].shape[0]
        df_summary.iat[27, 4] = "% no kill"
        if df_1E_2T_analysis['valid_2_contact'].shape[0] != 0:
            df_summary.iat[27, 5] = df_1E_2T_analysis['valid_2_contact_no_kill'].shape[0] / \
                df_1E_2T_analysis['valid_2_contact'].shape[0]

        for i, k in enumerate(['valid_1_contact', 'valid_1_contact_1_kill', 'valid_1_contact_no_kill']):  # 3
            df_summary.iat[24 + i, 8] = k
            df_summary.iat[24 + i, 9] = df_1E_2T_analysis[k].shape[0]
        df_summary.iat[27, 8] = "% no kill"
        if df_1E_2T_analysis['valid_1_contact'].shape[0] != 0:
            df_summary.iat[27, 9] = df_1E_2T_analysis['valid_1_contact_no_kill'].shape[0] / \
                df_1E_2T_analysis['valid_1_contact'].shape[0]

        for i, k in enumerate(['valid_no_contact', 'valid_no_contact_2_death', 'valid_no_contact_1_death', 'valid_no_contact_no_death']):
            df_summary.iat[24 + i, 12] = k
            df_summary.iat[24 + i, 13] = df_1E_2T_analysis[k].shape[0]
        df_summary.iat[27, 12] = "% no death"
        if df_1E_2T_analysis['valid_no_contact'].shape[0] != 0:
            df_summary.iat[27, 13] = df_1E_2T_analysis['valid_no_contact_no_death'].shape[0] / \
                df_1E_2T_analysis['valid_no_contact'].shape[0]

        df_1E_3T_analysis = {
            'valid': df_1E_3T_valid,
            'valid_3_contact': df_1E_3T_valid_3_contact,
            'valid_2_contact': df_1E_3T_valid_2_contact,
            'valid_1_contact': df_1E_3T_valid_1_contact,
            'valid_no_contact': df_1E_3T_valid_no_contact,
            'valid_p_contact_3_kill': df_1E_3T_valid_p_contact_3_kill,
            'valid_p_contact_2_kill': df_1E_3T_valid_p_contact_2_kill,
            'valid_p_contact_1_kill': df_1E_3T_valid_p_contact_1_kill,
            'valid_p_contact_no_kill': df_1E_3T_valid_p_contact_no_kill,
            'valid_no_contact_3_death': df_1E_3T_valid_no_contact_3_death,
            'valid_no_contact_2_death': df_1E_3T_valid_no_contact_2_death,
            'valid_no_contact_1_death': df_1E_3T_valid_no_contact_1_death,
            'valid_no_contact_no_death': df_1E_3T_valid_no_contact_no_death,
            'tot': df_1E_3T,
        }
        
        def get_valid_1E3T_rows(df):
            # Detect all tDeath columns
            t_death_cols = sorted(
                [col for col in df.columns if re.match(r"tDeath\d+", str(col))],
                key=lambda x: int(re.search(r"\d+", x).group())
            )
            
            # Valid: if all tDeath columns are NOT -2
            mask = pd.Series([True] * len(df), index=df.index)
            for col in t_death_cols:
                mask &= (df[col] != -2)
            
            return df[mask]

        def get_invalid_1E3T_rows(df):
            # Detect all tDeath columns
            t_death_cols = sorted(
                [col for col in df.columns if re.match(r"tDeath\d+", str(col))],
                key=lambda x: int(re.search(r"\d+", x).group())
            )
            
            # Invalid: if any tDeath column is -2
            mask = pd.Series([False] * len(df), index=df.index)
            for col in t_death_cols:
                mask |= (df[col] == -2)
            
            return df[mask]


     
        df_1E_3T_valid = get_valid_1E3T_rows(df_1E_3T)
        df_1E_3T_invalid = get_invalid_1E3T_rows(df_1E_3T)


        df_valid_3T = df_1E_3T_valid.copy()

        df_valid_3T["kill_count"] = (
            (df_valid_3T["tDeath1"] >= 0).astype(int) +
            (df_valid_3T["tDeath2"] >= 0).astype(int) +
            (df_valid_3T["tDeath3"] >= 0).astype(int)
        )

        df_valid_3T["has_contact"] = (
            (df_valid_3T["tContact1"] >= 0) |
            (df_valid_3T["tContact2"] >= 0) |
            (df_valid_3T["tContact3"] >= 0)
        )

        df_1E_3T_analysis = {
            'tot': df_1E_3T,
            'valid': df_valid_3T,
            'invalid': df_1E_3T_invalid,
            'valid_no_contact': df_valid_3T[df_valid_3T["has_contact"] == False],
            'valid_1_contact': df_valid_3T[
                (df_valid_3T[["tContact1", "tContact2", "tContact3"]] >= 0).sum(axis=1) == 1],
            'valid_2_contact': df_valid_3T[
                (df_valid_3T[["tContact1", "tContact2", "tContact3"]] >= 0).sum(axis=1) == 2],
            'valid_3_contact': df_valid_3T[
                (df_valid_3T[["tContact1", "tContact2", "tContact3"]] >= 0).sum(axis=1) == 3],
        }

        for i in range(4):
            key = f"valid_p_contact_{i}_kill"
            df_1E_3T_analysis[key] = df_valid_3T[(df_valid_3T["has_contact"]) & (df_valid_3T["kill_count"] == i)]

        for i in range(4):
            key = f"valid_no_contact_{i}_death"
            death_mask = (
                (df_valid_3T[["tContact1", "tContact2", "tContact3"]] < 0).all(axis=1) &
                ((df_valid_3T[["Absol tar death 1", "Absol tar death 2", "Absol tar death 3"]] >= 0).sum(axis=1) == i)
            )
            df_1E_3T_analysis[key] = df_valid_3T[death_mask]

        print("? 1E3T VALID wells:", df_1E_3T_analysis["valid"].shape[0])
        print("UIDs of valid 1E3T wells:", df_1E_3T_analysis["valid"]['UID'].tolist())

        row_start = 29
        df_summary.iat[row_start, 1] = '1E3T'
        df_summary.iat[row_start, 2] = df_1E_3T_analysis['tot'].shape[0]

        df_summary.iat[row_start + 1, 1] = '1E3T valid'
        df_summary.iat[row_start + 1, 2] = df_1E_3T_analysis['valid'].shape[0]
        df_summary.iat[row_start + 1, 3] = df_1E_3T_analysis['valid'].shape[0] / df_1E_3T_analysis['tot'].shape[0] if df_1E_3T_analysis['tot'].shape[0] else 0

        df_summary.iat[row_start + 2, 1] = '1E3T invalid'
        df_summary.iat[row_start + 2, 2] = df_1E_3T_analysis['invalid'].shape[0]
        df_summary.iat[row_start + 2, 3] = df_1E_3T_analysis['invalid'].shape[0] / df_1E_3T_analysis['tot'].shape[0] if df_1E_3T_analysis['tot'].shape[0] else 0

        df_summary.iat[row_start + 4, 2] = '1E3T no contact'
        df_summary.iat[row_start + 4, 3] = df_1E_3T_analysis['valid_no_contact'].shape[0]
        df_summary.iat[row_start + 5, 2] = '1E3T 1+ contact'
        df_summary.iat[row_start + 5, 3] = df_1E_3T_analysis['valid'].shape[0] - df_1E_3T_analysis['valid_no_contact'].shape[0]

        df_summary.iat[row_start + 7, 3] = '0 contacts'
        for i in range(4):
            label = f"{i} death"
            key = f"valid_no_contact_{i}_death"
            count = df_1E_3T_analysis[key].shape[0]
            df_summary.iat[row_start + 8 + i, 3] = label
            df_summary.iat[row_start + 8 + i, 4] = count
            df_summary.iat[row_start + 8 + i, 5] = count / df_1E_3T_analysis['valid_no_contact'].shape[0] if df_1E_3T_analysis['valid_no_contact'].shape[0] else 0

        df_summary.iat[row_start + 13, 3] = '1+ contact'
        contact_total = df_1E_3T_analysis['valid'].shape[0] - df_1E_3T_analysis['valid_no_contact'].shape[0]
        for i in range(4):
            label = f"{i} kill"
            key = f"valid_p_contact_{i}_kill"
            count = df_1E_3T_analysis[key].shape[0]
            df_summary.iat[row_start + 14 + i, 3] = label
            df_summary.iat[row_start + 14 + i, 4] = count
            df_summary.iat[row_start + 14 + i, 5] = count / contact_total if contact_total else 0

        print("?? Kill count breakdown (1E3T valid wells):")
        print(df_valid_3T["kill_count"].value_counts().sort_index())

        print("?? Contact status breakdown:")
        print(df_valid_3T["has_contact"].value_counts())


        # if eff_present:
        #     kills = [int(entry_info[self.result_field_inds[t_death]] > 0 and
        #                  entry_info[self.result_field_inds[t_contact]] > 0) for t_death, t_contact in
        #              [["tDeath1", "tContact1"],
        #               ["tDeath2", "tContact2"],
        #               ["tDeath3", "tContact3"],
        #               ["tDeath4", "tContact4"],
        #               ["tDeath5", "tContact5"],]]
        #     kill_count = sum(kills[:no_targets])

        #     deaths = [int(entry_info[self.result_field_inds[t_death]] != -1 and
        #                   entry_info[self.result_field_inds[t_contact]] != -2) for t_death, t_contact in
        #               [["tDeath1", "tContact1"],
        #               ["tDeath2", "tContact2"],
        #               ["tDeath3", "tContact3"],
        #               ["tDeath4", "tContact4"],
        #               ["tDeath5", "tContact5"],]]
        #     death_count = sum(deaths[:no_targets])

        #     kill_counter[no_targets][kill_count] += 1
        #     death_counter[no_targets][death_count] += 1

        # df_0E_1T = create_0E_1T(df, LastColumn, Wells)
        # df_1E_0T = create_1E_0T(df, LastColumn, Wells)
        # df_1E_1T = create_1E_XT(df, LastColumn, Wells, 1)
        # df_1E_2T = create_1E_XT(df, LastColumn, Wells, 2)
        # df_1E_3T = create_1E_XT(df, LastColumn, Wells, 3)
        # df_1E_PT = create_1E_posT(df, LastColumn, Wells)
        with pd.ExcelWriter(os.path.join(home_directory, "event_breakdown.xlsx")) as writer:

            # Write the Master sheet
            self.df.to_excel(writer, sheet_name="Master", index=False)
            
            # Helper function to add columns
            def add_final_minute_enable(df, col_name):
                df = df.copy()
                if col_name in df.columns:
                    df["final_minute"] = df[col_name].where(df[col_name] != -1, 360)
                    df["Enable"] = (df[col_name] != -1).astype(int)
                    print(f"Added final_minute and Enable columns using {col_name}")
                else:
                    print(f"Column {col_name} not found in DataFrame. Available columns: {list(df.columns)}")
                return df

            # Write Table_1E1+T
            df_1E_PT.to_excel(writer, sheet_name="Table_1E1+T", index=False)

            # Debug: Print available keys
            print("0E_1T analysis keys:", list(df_0E_1T_analysis.keys()))
            print("1E_0T analysis keys:", list(df_1E_0T_analysis.keys()))
            print("1E_1T analysis keys:", list(df_1E_1T_analysis.keys()))

            # Write 0E_1T sheets
            for k, v in df_0E_1T_analysis.items():
                k_name = k.replace("contact", "c").replace("valid", "v")
                sheet_name = f"0E_1T_{k_name} (n={v.shape[0]})"
                if k == "tot" and "TarDeath" in v.columns:
                    v = add_final_minute_enable(v, "TarDeath")
                v.to_excel(writer, sheet_name=sheet_name, index=False)

            # Write 1E_0T sheets
            for k, v in df_1E_0T_analysis.items():
                k_name = k.replace("contact", "c").replace("valid", "v")
                sheet_name = f"1E_0T_{k_name} (n={v.shape[0]})"
                if k == "tot" and "tEffDeath" in v.columns:
                    v = add_final_minute_enable(v, "tEffDeath")
                v.to_excel(writer, sheet_name=sheet_name, index=False)

            # Write 1E_1T sheets
            for k, v in df_1E_1T_analysis.items():
                k_name = k.replace("contact", "c").replace("valid", "v")
                sheet_name = f"1E_1T_{k_name} (n={v.shape[0]})"
                if k == "tot" and "tDeath1" in v.columns:
                    v = add_final_minute_enable(v, "tDeath1")
                v.to_excel(writer, sheet_name=sheet_name, index=False)

            # Write 1E_2T sheets
            for k, v in df_1E_2T_analysis.items():
                k_name = k.replace("contact", "c").replace("valid", "v")
                v.to_excel(writer, sheet_name=f"1E_2T_{k_name} (n={v.shape[0]})", index=False)
            # Write 1E_3T sheets
            for k, v in df_1E_3T_analysis.items():
                k_name = k.replace("contact", "c").replace("valid", "v")
                v.to_excel(writer, sheet_name=f"1E_3T_{k_name} (n={v.shape[0]})", index=False)
            # Write summary
            df_summary.to_excel(writer, sheet_name="Summary", index=False)

            # Create '1E_1T kill curve' sheet if possible
            # Use the 'tot' DataFrame from df_1E_1T_analysis
            tot_df = df_1E_1T_analysis.get("tot")
            if tot_df is not None and "tDeath1" in tot_df.columns:
                # Add the columns first
                kill_curve_df = add_final_minute_enable(tot_df, "tDeath1")
                # Select only the required columns
                kill_curve_df = kill_curve_df[["UID", "tDeath1", "final_minute", "Enable"]].copy()
                kill_curve_df.to_excel(writer, sheet_name="1E_1T kill curve", index=False)

        with pd.ExcelWriter(os.path.join(home_directory, "data_summary.xlsx")) as writer:

            df_1E_PT.to_excel(writer, sheet_name="Table_1E1+T", index=False)
            df_0E_1T_analysis['tot'].to_excel(
                writer, sheet_name=f"0E_1T (n={df_0E_1T_analysis['tot'].shape[0]})", index=False)
            df_1E_0T_analysis['tot'].to_excel(
                writer, sheet_name=f"1E_0T (n={df_1E_0T_analysis['tot'].shape[0]})", index=False)
            df_1E_1T_analysis['tot'].to_excel(
                writer, sheet_name=f"1E_1T (n={df_1E_1T_analysis['tot'].shape[0]})", index=False)
            df_1E_2T_analysis['tot'].to_excel(
                writer, sheet_name=f"1E_2T (n={df_1E_2T_analysis['tot'].shape[0]})", index=False)
            df_1E_3T_analysis['tot'].to_excel(
                writer, sheet_name=f"1E_3T (n={df_1E_3T_analysis['tot'].shape[0]})", index=False)
            df_summary.to_excel(writer, sheet_name="Summary", index=False)

        df_1E_0T.to_csv(filename_1E_0T, index=False)
        df_0E_1T.to_csv(filename_0E_1T, index=False)
        df_1E_1T.to_csv(filename_1E_1T, index=False)
        df_1E_2T.to_csv(filename_1E_2T, index=False)
        df_1E_3T.to_csv(filename_1E_3T, index=False)
        df_1E_PT.to_csv(filename_1E_PT, index=False)

        #print(death_counter)
        #print(kill_counter)

        # Add at the end of the secondary() method, after saving data_summary.xlsx:
        def format_summary_excel(output_path):
            wb = openpyxl.load_workbook(output_path)
            ws = wb['Summary']

            yellow = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
            light_grey = PatternFill(start_color='DDDDDD', end_color='DDDDDD', fill_type='solid')
            light_purple = PatternFill(start_color='FFCCFF', end_color='FFCCFF', fill_type='solid')
            bold_font = Font(bold=True)

            def write_block(ws_out, start_row, start_col, block, header=False, header_color=None, value_grey=True):
                for i, row in enumerate(block):
                    for j, val in enumerate(row):
                        cell = ws_out.cell(row=start_row + i, column=start_col + j, value=val)
                        if header and i == 0:
                            cell.font = bold_font
                            if header_color:
                                cell.fill = header_color
                        elif value_grey and isinstance(val, int):
                            cell.fill = light_grey

            if 'FormattedSummary' in wb.sheetnames:
                del wb['FormattedSummary']
            ws_out = wb.create_sheet('FormattedSummary')

            # Helper: Find the row index of a section header
            def find_section_row(section_label):
                for idx, row in enumerate(ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=2, max_col=2), 1):
                    if row[0].value and section_label.lower() in str(row[0].value).strip().lower():
                        return idx
                return None

            # Helper: Find the value for a label within a section (robust, partial/case-insensitive match)
            def find_value_in_section(section_label, label, max_rows=20, col=2, exact=True):
                start_row = find_section_row(section_label)
                if start_row is None:
                    return 0
                # Only look in the specified column (default col=2, i.e., column B)
                for row in ws.iter_rows(min_row=start_row+1, max_row=min(start_row+max_rows, ws.max_row), min_col=col, max_col=col):
                    cell_value = str(row[0].value).strip().lower() if row[0].value else ""
                    if (exact and label.lower() == cell_value) or (not exact and label.lower() in cell_value):
                        # Get the value from the next column (col+1)
                        return ws.cell(row=row[0].row, column=col+1).value
                return 0

            # Helper: For contact breakdowns, search all columns and use partial match
            def find_value_any_col(section_label, label, max_rows=20):
                start_row = find_section_row(section_label)
                if start_row is None:
                    return 0
                for row in ws.iter_rows(min_row=start_row+1, max_row=min(start_row+max_rows, ws.max_row), min_col=2, max_col=ws.max_column):
                    for i, cell in enumerate(row):
                        cell_value = str(cell.value).strip().lower() if cell.value else ""
                        if label.lower() in cell_value:
                            if i + 1 < len(row):
                                return row[i + 1].value
                return 0

            row_ptr = 1
            # --- 0E 1T Section ---
            ws_out.cell(row=row_ptr, column=2, value='0E 1T').font = bold_font
            ws_out.cell(row=row_ptr, column=2).fill = yellow
            row_ptr += 1

            # Find the first sheet containing '0E_1T'
            sheet_0e1t = next((s for s in wb.sheetnames if '0E_1T' in s), None)
            if sheet_0e1t is None:
                raise ValueError("No 0E_1T sheet found in workbook!")
            df_0e1t = pd.read_excel(output_path, sheet_name=sheet_0e1t)

            tot = len(df_0e1t)
            invalid_0e1t = ((df_0e1t['TarDeath'] == 0) | (df_0e1t['TarDeath'] == 5)).sum()
            valid_0e1t = tot - invalid_0e1t
            alive_0e1t = (df_0e1t['TarDeath'] == -1).sum()
            dead_0e1t = valid_0e1t - alive_0e1t

            # --- Overwrite Summary sheet values for 0E 1T ---
            section_label = '0E 1T'
            start_row = find_section_row(section_label)
            if start_row:
                # Overwrite values for valid, invalid, alive, dead
                for label, value in [
                    ('valid', valid_0e1t),
                    ('invalid', invalid_0e1t),
                    ('alive', alive_0e1t),
                    ('dead', dead_0e1t)
                ]:
                    for row in ws.iter_rows(min_row=start_row+1, max_row=min(start_row+20, ws.max_row), min_col=2, max_col=2):
                        if row[0].value and str(row[0].value).strip().lower() == label:
                            ws.cell(row=row[0].row, column=3, value=value)
                            break

            valid_fraction = round(valid_0e1t / tot, 4) if tot else 0
            invalid_fraction = round(invalid_0e1t / tot, 4) if tot else 0
            alive_fraction = round(alive_0e1t / tot, 4) if tot else 0
            dead_fraction = round(dead_0e1t / tot, 4) if tot else 0
            block = [
                ['total', tot, '', ''],
                ['valid', valid_0e1t, valid_fraction, ''],
                ['invalid', invalid_0e1t, invalid_fraction, ''],
                ['', 'dead', dead_0e1t, dead_fraction],
                ['', 'alive', alive_0e1t, alive_fraction]
            ]
            write_block(ws_out, row_ptr, 2, block, value_grey=True)
            row_ptr += len(block) + 1

            # --- 1E 0T Section ---
            ws_out.cell(row=row_ptr, column=2, value='1E 0T').font = bold_font
            ws_out.cell(row=row_ptr, column=2).fill = yellow
            row_ptr += 1

            # Find the first sheet containing '1E_0T'
            sheet_1e0t = next((s for s in wb.sheetnames if '1E_0T' in s), None)
            if sheet_1e0t is None:
                raise ValueError("No 1E_0T sheet found in workbook!")
            df_1e0t = pd.read_excel(output_path, sheet_name=sheet_1e0t)

            tot = len(df_1e0t)
            invalid_1e0t = ((df_1e0t['tEffDeath'] == 0) | (df_1e0t['tEffDeath'] == 5)).sum()
            valid_1e0t = tot - invalid_1e0t
            alive_1e0t = (df_1e0t['tEffDeath'] == -1).sum()
            dead_1e0t = valid_1e0t - alive_1e0t

            # --- Overwrite Summary sheet values for 1E 0T ---
            section_label = '1E 0T'
            start_row = find_section_row(section_label)
            if start_row:
                # Overwrite values for valid, invalid, alive, dead
                for label, value in [
                    ('valid', valid_1e0t),
                    ('invalid', invalid_1e0t),
                    ('alive', alive_1e0t),
                    ('dead', dead_1e0t)
                ]:
                    for row in ws.iter_rows(min_row=start_row+1, max_row=min(start_row+20, ws.max_row), min_col=2, max_col=2):
                        if row[0].value and str(row[0].value).strip().lower() == label:
                            ws.cell(row=row[0].row, column=3, value=value)
                            break

            valid_fraction = round(valid_1e0t / tot, 4) if tot else 0
            invalid_fraction = round(invalid_1e0t / tot, 4) if tot else 0
            alive_fraction = round(alive_1e0t / tot, 4) if tot else 0
            dead_fraction = round(dead_1e0t / tot, 4) if tot else 0
            block = [
                ['total', tot, '', ''],
                ['valid', valid_1e0t, valid_fraction, ''],
                ['invalid', invalid_1e0t, invalid_fraction, ''],
                ['', 'dead', dead_1e0t, dead_fraction],
                ['', 'alive', alive_1e0t, alive_fraction]
            ]
            write_block(ws_out, row_ptr, 2, block, value_grey=True)
            row_ptr += len(block) + 1

            # --- 1E 1T Section ---
            ws_out.cell(row=row_ptr, column=2, value='1E 1T').font = bold_font
            ws_out.cell(row=row_ptr, column=2).fill = yellow
            row_ptr += 1
            tot = find_value_in_section('1E 1T', 'tot', col=2, exact=True)
            valid = find_value_in_section('1E 1T', 'valid', col=2, exact=True)
            invalid = tot - valid if tot is not None and valid is not None else None
            valid_fraction = round(valid / tot, 4) if tot else 0
            invalid_fraction = round(invalid / tot, 4) if tot else 0
            block = [
                ['total', tot, '', ''],
                ['valid', valid, valid_fraction],
                ['invalid', invalid, invalid_fraction]
            ]
            write_block(ws_out, row_ptr, 2, block, value_grey=True)
            row_ptr += len(block)
            # Contact breakdown (use flexible lookup)
            valid_no_contact = find_value_any_col('1E 1T', 'valid_no_contact') or 0
            valid_contact = find_value_any_col('1E 1T', 'valid_contact') or 0
            fraction_0_contact = round(valid_no_contact / valid, 4) if valid else 0
            fraction_1_contact = round(valid_contact / valid, 4) if valid else 0
            block = [
                ['', '0 contact', valid_no_contact, fraction_0_contact],
                ['', '1 contact', valid_contact, fraction_1_contact]
            ]
            write_block(ws_out, row_ptr, 2, block, value_grey=True)
            row_ptr += len(block)
            # Contact constrained killing (side by side)
            ws_out.merge_cells(start_row=row_ptr, start_column=5, end_row=row_ptr, end_column=12)
            cell = ws_out.cell(row=row_ptr, column=5)
            cell.value = 'Contact constrained killing'
            cell.font = bold_font
            cell.fill = light_purple
            cell.alignment = Alignment(horizontal='center', vertical='center')
            row_ptr += 1
            # Prepare 0 contact block
            valid_no_contact_no_death = find_value_any_col('1E 1T', 'valid_no_contact_no_death')
            valid_no_contact_death = find_value_any_col('1E 1T', 'valid_no_contact_death')
            fraction_alive = round(valid_no_contact_no_death / valid_no_contact, 4) if valid_no_contact else 0
            fraction_dead = round(valid_no_contact_death / valid_no_contact, 4) if valid_no_contact else 0
            block_0 = [
                ['0 contact', '', '', ''],
                ['total', valid_no_contact, '', ''],
                ['0 dead', valid_no_contact_no_death, fraction_alive, ''],
                ['1 dead', valid_no_contact_death, fraction_dead, '']
            ]
            # Prepare 1 contact block
            valid_contact_no_kill = find_value_any_col('1E 1T', 'valid_contact_no_kill')
            valid_contact_kill = find_value_any_col('1E 1T', 'valid_contact_kill')
            fraction_0_killed = round(valid_contact_no_kill / valid_contact, 4) if valid_contact else 0
            fraction_1_killed = round(valid_contact_kill / valid_contact, 4) if valid_contact else 0
            block_1 = [
                ['1 contact', '', '', ''],
                ['total', valid_contact, '', ''],
                ['0 killed', valid_contact_no_kill, fraction_0_killed, ''],
                ['1 killed', valid_contact_kill, fraction_1_killed, '']
            ]
            # Write both blocks side by side
            start_row = row_ptr
            col_0 = 5
            col_1 = 9
            write_block(ws_out, start_row, col_0, block_0, value_grey=True)
            write_block(ws_out, start_row, col_1, block_1, value_grey=True)
            row_ptr += max(len(block_0), len(block_1))

            # --- 1E 2T Section ---
            ws_out.cell(row=row_ptr, column=2, value='1E 2T').font = bold_font
            ws_out.cell(row=row_ptr, column=2).fill = yellow
            row_ptr += 1
            tot = find_value_in_section('1E 2T', 'tot', col=2, exact=True)
            valid = find_value_in_section('1E 2T', 'valid', col=2, exact=True)
            invalid = tot - valid if tot is not None and valid is not None else None
            valid_fraction = round(valid / tot, 4) if tot else 0
            invalid_fraction = round(invalid / tot, 4) if tot else 0
            block = [
                ['total', tot, '', ''],
                ['valid', valid, valid_fraction],
                ['invalid', invalid, invalid_fraction]
            ]
            write_block(ws_out, row_ptr, 2, block, value_grey=True)
            row_ptr += len(block)
            # Contact breakdown (use flexible lookup)
            valid_no_contact = find_value_any_col('1E 2T', 'valid_no_contact')
            valid_1_contact = find_value_any_col('1E 2T', 'valid_1_contact')
            valid_2_contact = find_value_any_col('1E 2T', 'valid_2_contact')
            valid_1plus_contact = valid_1_contact + valid_2_contact if valid_1_contact and valid_2_contact else 0
            fraction_0_contact = round(valid_no_contact / valid, 4) if valid else 0
            fraction_1_contact = round(valid_1_contact / valid, 4) if valid else 0
            fraction_2_contact = round(valid_2_contact / valid, 4) if valid else 0
            fraction_1plus_contact = round(valid_1plus_contact / valid, 4) if valid else 0
            block = [
                ['', '0 contact', valid_no_contact, fraction_0_contact],
                ['', '1 contact', valid_1_contact, fraction_1_contact],
                ['', '2 contact', valid_2_contact, fraction_2_contact],
                ['', '1+ contact', valid_1plus_contact, fraction_1plus_contact]
            ]
            write_block(ws_out, row_ptr, 2, block, value_grey=True)
            row_ptr += len(block)
            # Contact constrained killing (side by side)
            ws_out.merge_cells(start_row=row_ptr, start_column=5, end_row=row_ptr, end_column=16)
            cell = ws_out.cell(row=row_ptr, column=5)
            cell.value = 'Contact constrained killing'
            cell.font = bold_font
            cell.fill = light_purple
            cell.alignment = Alignment(horizontal='center', vertical='center')
            row_ptr += 1
            # 0 contact block
            valid_no_contact_0_killed = find_value_any_col('1E 2T', 'valid_no_contact')
            valid_no_contact_1_killed = find_value_any_col('1E 2T', 'valid_no_contact_1_death')
            valid_no_contact_2_killed = find_value_any_col('1E 2T', 'valid_no_contact_2_death')
            fraction_0_killed = round(valid_no_contact_0_killed / valid_no_contact, 4) if valid_no_contact else 0
            fraction_1_killed = round(valid_no_contact_1_killed / valid_no_contact, 4) if valid_no_contact else 0
            fraction_2_killed = round(valid_no_contact_2_killed / valid_no_contact, 4) if valid_no_contact else 0
            block_0 = [
                ['0 contact', '', '', ''],
                ['total', valid_no_contact, '', ''],
                ['0 killed', valid_no_contact_0_killed, fraction_0_killed, ''],
                ['1 killed', valid_no_contact_1_killed, fraction_1_killed, ''],
                ['2 killed', valid_no_contact_2_killed, fraction_2_killed, '']
            ]
            # 1 contact block
            valid_1_contact_no_kill = find_value_any_col('1E 2T', 'valid_1_contact_no_kill')
            valid_1_contact_1_kill = find_value_any_col('1E 2T', 'valid_1_contact_1_kill')
            fraction_0_killed = round(valid_1_contact_no_kill / valid_1_contact, 4) if valid_1_contact else 0
            fraction_1_killed = round(valid_1_contact_1_kill / valid_1_contact, 4) if valid_1_contact else 0
            block_1 = [
                ['1 contact', '', '', ''],
                ['total', valid_1_contact, '', ''],
                ['0 killed', valid_1_contact_no_kill, fraction_0_killed, ''],
                ['1 killed', valid_1_contact_1_kill, fraction_1_killed, ''],
                ['2 killed', 0, 0, '']
            ]
            # 2 contact block
            valid_2_contact_1_kill = find_value_any_col('1E 2T', 'valid_2_contact_1_kill')
            valid_2_contact_2_kill = find_value_any_col('1E 2T', 'valid_2_contact_2_kill')
            valid_2_contact_0_killed = valid_2_contact - valid_2_contact_1_kill - valid_2_contact_2_kill if valid_2_contact and valid_2_contact_1_kill and valid_2_contact_2_kill else 0
            fraction_0_killed = round(valid_2_contact_0_killed / valid_2_contact, 4) if valid_2_contact else 0
            fraction_1_killed = round(valid_2_contact_1_kill / valid_2_contact, 4) if valid_2_contact else 0
            fraction_2_killed = round(valid_2_contact_2_kill / valid_2_contact, 4) if valid_2_contact else 0
            block_2 = [
                ['2 contact', '', '', ''],
                ['total', valid_2_contact, '', ''],
                ['0 killed', valid_2_contact_0_killed, fraction_0_killed, ''],
                ['1 killed', valid_2_contact_1_kill, fraction_1_killed, ''],
                ['2 killed', valid_2_contact_2_kill, fraction_2_killed, '']
            ]
            # Write all blocks side by side
            start_row = row_ptr
            col_0 = 5
            col_1 = 9
            col_2 = 13
            write_block(ws_out, start_row, col_0, block_0, value_grey=True)
            write_block(ws_out, start_row, col_1, block_1, value_grey=True)
            write_block(ws_out, start_row, col_2, block_2, value_grey=True)
            row_ptr += max(len(block_0), len(block_1), len(block_2))

            wb.save(output_path)

        # After saving data_summary.xlsx in secondary():
        output_path = os.path.join(home_directory, "data_summary.xlsx")
        format_summary_excel(output_path)

        event_breakdown_path = os.path.join(home_directory, "event_breakdown.xlsx")
        format_summary_excel(event_breakdown_path)

    def format_0E(self, df):
        cols = [
            "UID", "Tar1 dWell No Cont (um)", "Tar1 AR No Cont", "Absol tar death 1"]
        new_df = df[cols].copy()
        new_df.rename(columns={'Absol tar death 1': 'TarDeath'}, inplace=True)
        return new_df

    def format_0T(self, df):
        cols = [
            "UID", "Eff dWell Cont (um)", "Eff AR No Cont", "tEffDeath"]
        new_df = df[cols].copy()
        new_df.rename(
            columns={'Eff dWell Cont (um)': 'dWell No Cont(um)', "Eff AR No Cont": "AR No Cont"}, inplace=True)
        return new_df

    def format_1E(self, df, t_val):
        cols = ["UID", "No of targets"]
        for i in range(t_val):
            cols.extend([f"tSeek{i+1}", f"tDeath{i+1}", f"tContact{i+1}",
                        f"Cum contact {i+1}", f"Absol tar death {i+1}"])
        cols.extend(["Eff dWell No Cont (um)", "Eff dWell Cont (um)",
                    "Eff AR No Cont", "Eff AR Cont", "tEffDeath"])
        for i in range(t_val):
            cols.extend([f"Tar{i+1} dWell No Cont (um)", f"Tar{i+1} dWell Cont (um)",
                        f"Tar{i+1} AR No Cont", f"Tar{i+1} AR Cont"])
        new_df = df[cols].copy()
        return new_df


def run_TIMING_secondary(path, DInt, EDInt, ContInt, TimInt, Pix):
    labels = ["Eff_x", "Eff_y", "Eff_AR", "Eff_Speed", "Eff_death_int", "Eff_death", "Tar1x", "Tar1y", "Target1_AR",
              "Tar1_Speed", "Tar1D_Int", "Tar1D", "Cont1_Int", "Cont1", "Tar2x", "Tar2y", "Target2_AR", "Tar2_Speed", "Tar2D_Int",
              "Tar2D", "Cont2_Int", "Cont2", "Tar3x", "Tar3y", "Target3_AR", "Tar3_Speed", "Tar3D_Int", "Tar3D", "Cont3_Int", "Cont3",
              "Tar4x", "Tar4y", "Target4_AR", "Tar4_Speed", "Tar4D_Int", "Tar4D", "Cont4_Int", "Cont4", "Tar5x", "Tar5y", "Target5_AR",
              "Tar5_Speed", "Tar5D_Int", "Tar5D", "Cont5_Int", "Cont5", "BEAD_I", "BEAD_II", 'Time_Point']

    result_fields = [
        "UID", "No of targets",
        "tSeek1", "tDeath1", "tContact1", "Cum contact 1", "Absol tar death 1",
        "tSeek2", "tDeath2", "tContact2", "Cum contact 2", "Absol tar death 2",
        "tSeek3", "tDeath3", "tContact3", "Cum contact 3", "Absol tar death 3",
        "tSeek4", "tDeath4", "tContact4", "Cum contact 4", "Absol tar death 4",
        "tSeek5", "tDeath5", "tContact5", "Cum contact 5", "Absol tar death 5",
        "Eff dWell No Cont (um)", "Eff dWell Cont (um)", "Eff AR No Cont", "Eff AR Cont", "tEffDeath",
        "Tar1 dWell No Cont (um)", "Tar1 dWell Cont (um)", "Tar1 AR No Cont", "Tar1 AR Cont",
        "Tar2 dWell No Cont (um)", "Tar2 dWell Cont (um)", "Tar2 AR No Cont", "Tar2 AR Cont",
        "Tar3 dWell No Cont (um)", "Tar3 dWell Cont (um)", "Tar3 AR No Cont", "Tar3 AR Cont",
        "Tar4 dWell No Cont (um)", "Tar4 dWell Cont (um)", "Tar4 AR No Cont", "Tar4 AR Cont",
        "Tar5 dWell No Cont (um)", "Tar5 dWell Cont (um)", "Tar5 AR No Cont", "Tar5 AR Cont"
    ]
    sec = TIMING_secondary(
        path=path, result_fields=result_fields, labels=labels)
    sec.secondary(DInt, EDInt, ContInt, TimInt, Pix)


def create_0E_1T(df, lc, w_num):
    # Create an empty DataFrame
    df_rt = pd.DataFrame(
        columns=["UID", "Tar dWell No Cont(um)", "Tar AR No Cont", "TarDeath"])

    # Populate the DataFrame
    K = 0
    for i in range(1, w_num + 1):
        if df.iloc[(i - 1) * 49, lc + 2] == 1 and df.iloc[(i - 1) * 49, lc] == 0:
            uid = df.iloc[(i - 1) * 49, 0]
            tar_dwell = df.iloc[(i - 1) * 49, lc + 12]
            tar_ar = df.iloc[(i - 1) * 49, lc + 14]
            tar_death = df.iloc[(i - 1) * 49, lc + 6]

            df_rt.loc[K] = [uid, tar_dwell, tar_ar, tar_death]
            K += 1

    return df_rt


def create_1E_0T(df, lc, w_num):
    # Create an empty DataFrame
    df_rt = pd.DataFrame(
        columns=["UID", "dWell No Cont(um)", "AR No Cont", "tEffDeath"])

    # Populate the DataFrame based on your original logic
    K = 0
    for i in range(1, w_num + 1):
        if df.iloc[(i - 1) * 49, lc] == 1 and df.iloc[(i - 1) * 49, lc + 2] == 0:
            df_rt.loc[K] = [
                df.iloc[(i - 1) * 49, 0],
                df.iloc[(i - 1) * 49, lc + 8],
                df.iloc[(i - 1) * 49, lc + 10],
                df.iloc[(i - 1) * 49, lc + 4]
            ]
            K += 1

    return df_rt


def create_1E_XT(df, lc, w_num, X):
    # Create a DataFrame to hold the data
    df_rt = pd.DataFrame(columns=[
        "UID", "No of targets",
        "tSeek1", "tDeath1", "tContact1", "Cum contact 1", "Absol tar death 1",
        "tSeek2", "tDeath2", "tContact2", "Cum contact 2", "Absol tar death 2",
        "tSeek3", "tDeath3", "tContact3", "Cum contact 3", "Absol tar death 3",
        "tSeek4", "tDeath4", "tContact4", "Cum contact 4", "Absol tar death 4",
        "tSeek5", "tDeath5", "tContact5", "Cum contact 5", "Absol tar death 5",
        "Eff dWell No Cont (um)", "Eff dWell Cont (um)", "Eff AR No Cont", "Eff AR Cont", "tEffDeath",
        "Tar1 dWell No Cont (um)", "Tar1 dWell Cont (um)", "Tar1 AR No Cont", "Tar1 AR Cont",
        "Tar2 dWell No Cont (um)", "Tar2 dWell Cont (um)", "Tar2 AR No Cont", "Tar2 AR Cont",
        "Tar3 dWell No Cont (um)", "Tar3 dWell Cont (um)", "Tar3 AR No Cont", "Tar3 AR Cont",
        "Tar4 dWell No Cont (um)", "Tar4 dWell Cont (um)", "Tar4 AR No Cont", "Tar4 AR Cont",
        "Tar5 dWell No Cont (um)", "Tar5 dWell Cont (um)", "Tar5 AR No Cont", "Tar5 AR Cont"
    ])
    all_tup = [(0, 3), (0, 6), (0, 7), (0, 16), (0, 5), (8, 3), (8, 6), (8, 7), (8, 16), (8, 5), (16, 3), (16, 6), (16, 7), (16, 16), (16, 5), (24, 3), (24, 6), (24, 7), (24, 16), (24, 5), (32, 3), (32, 6), (32, 7), (32, 16), (32, 5),
               (0, 8), (0, 9), (0, 10), (0, 11), (0, 4), (0, 12), (0, 13), (0, 14), (0, 15), (8, 12), (8, 13), (8, 14), (8, 15), (16, 12), (16, 13), (16, 14), (16, 15), (24, 12), (24, 13), (24, 14), (24, 15), (32, 12), (32, 13), (32, 14), (32, 15)]
    K = 0
    for i in range(1, w_num + 1):
        if df.iloc[(i - 1) * 49, lc] == 1 and df.iloc[(i - 1) * 49, lc + 2] == X:
            print("ABAAABBABA")
            print((i - 1) * 49 + (52 - 33) * 8 + 1)
            df_rt.loc[K] = [df.iloc[(i - 1) * 49, 0], df.iloc[(i - 1) * 49, lc + 2]] \
                + [df.iloc[(i - 1) * 49 + a, lc + b] for a, b in all_tup]
            # + [df.iloc[(i - 1) * 49 + (j - 4) * 8 + 1, lc + (j - 1)] for j in range(4, 33)] \
            # + [df.iloc[(i - 1) * 49 + (j - 33) * 8 + 1, lc + (j - 29)] for j in range(33,53)]

            K += 1

    return df_rt


def create_1E_posT(df, lc, w_num):
    # Create a DataFrame to hold the data
    df_rt = pd.DataFrame(columns=[
        "UID", "No of targets",
        "tSeek1", "tDeath1", "tContact1", "Cum contact 1", "Absol tar death 1",
        "tSeek2", "tDeath2", "tContact2", "Cum contact 2", "Absol tar death 2",
        "tSeek3", "tDeath3", "tContact3", "Cum contact 3", "Absol tar death 3",
        "tSeek4", "tDeath4", "tContact4", "Cum contact 4", "Absol tar death 4",
        "tSeek5", "tDeath5", "tContact5", "Cum contact 5", "Absol tar death 5",
        "Eff dWell No Cont (um)", "Eff dWell Cont (um)", "Eff AR No Cont", "Eff AR Cont", "tEffDeath",
        "Tar1 dWell No Cont (um)", "Tar1 dWell Cont (um)", "Tar1 AR No Cont", "Tar1 AR Cont",
        "Tar2 dWell No Cont (um)", "Tar2 dWell Cont (um)", "Tar2 AR No Cont", "Tar2 AR Cont",
        "Tar3 dWell No Cont (um)", "Tar3 dWell Cont (um)", "Tar3 AR No Cont", "Tar3 AR Cont",
        "Tar4 dWell No Cont (um)", "Tar4 dWell Cont (um)", "Tar4 AR No Cont", "Tar4 AR Cont",
        "Tar5 dWell No Cont (um)", "Tar5 dWell Cont (um)", "Tar5 AR No Cont", "Tar5 AR Cont"
    ])
    """
        UID, [0,actual 0]
        No of targets, [0,2]
       X tSeek1, [0,3] 
       X tDeath1, [0,6]
        tContact1, [0,7]
       X Cum contact 1, [0,16]
       X Absol tar death 1, [0,5]
       X tSeek2, [8,3]
       X tDeath2, [8,6]
        tContact2, [8,7]
       X Cum contact 2, [8,16]
       X Absol tar death 2, [8,5]
       X tSeek3, [16,3]
       X tDeath3, [16,6]
        tContact3, [16,7]
       X Cum contact 3, [16,16]
       X Absol tar death 3, [16,5]
       X tSeek4, [24,3]
       X tDeath4, [24,6]
        tContact4, [24,7]
       X Cum contact 4, [24,16]
       X Absol tar death 4, [24,5]
       X tSeek5, [32,3]
       X tDeath5, [32,6]
        tContact5, [32,7]
       X Cum contact 5, [32,16]
       X Absol tar death 5, [32,5]
        Eff dWell No Cont (um), [0,8]
        Eff dWell Cont (um), [0,9]
        Eff AR No Cont, [0,10]
        Eff AR Cont, [0,11]
       X tEffDeath, [0,4]
        Tar1 dWell No Cont (um), [0,12]
        Tar1 dWell Cont (um), [0,13]
        Tar1 AR No Cont, [0,14]
        Tar1 AR Cont, [0,15]
        Tar2 dWell No Cont (um), [8,12]
        Tar2 dWell Cont (um), [8,13]
        Tar2 AR No Cont, [8,14]
        Tar2 AR Cont, [8,15]
        Tar3 dWell No Cont (um), [16,12]
        Tar3 dWell Cont (um), [16,13]
        Tar3 AR No Cont, [16,14]
        Tar3 AR Cont, [16,15]
        Tar4 dWell No Cont (um), [24,12]
        Tar4 dWell Cont (um), [24,13]
        Tar4 AR No Cont, [24,14]
        Tar4 AR Cont, [24,15]
        Tar5 dWell No Cont (um), [32,12]
        Tar5 dWell Cont (um), [32,13]
        Tar5 AR No Cont, [32,14]
        Tar5 AR Cont, [32,15]
    """
    all_tup = [(1, 3), (1, 6), (1, 7), (1, 16), (1, 5), (9, 3), (9, 6), (9, 7), (9, 16), (9, 5), (17, 3), (17, 6), (17, 7),
               (17, 16), (17, 5), (25, 3), (25, 6), (25, 7), (25,
                                                              16), (25, 5), (33, 3), (33, 6), (33, 7), (33, 16), (33, 5),
               (1, 8), (1, 9), (1, 10), (1, 11), (1, 4), (1, 12), (1,
                                                                   13), (1, 14), (1, 15), (9, 12), (9, 13), (9, 14), (9, 15),
               (17, 12), (17, 13), (17, 14), (17, 15), (25, 12), (25,
                                                                  13), (25, 14), (25, 15), (33, 12), (33, 13),
               (33, 14), (33, 15)]
    K = 0
    for i in range(1, w_num + 1):
        if df.iloc[(i - 1) * 49, lc] == 1 and df.iloc[(i - 1) * 49, lc + 2] > 0:
            print("ABAAABBABA")
            print((i - 1) * 49 + (52 - 33) * 8 + 1)
            df_rt.loc[K] = [df.iloc[(i - 1) * 49, 0], df.iloc[(i - 1) * 49, lc + 2]] \
                + [df.iloc[(i - 1) * 49 + a - 1, lc + b] for a, b in all_tup]
            # + [df.iloc[(i - 1) * 49 + (j - 4) * 8 + 1, lc + (j - 1)] for j in range(4, 33)] \
            # + [df.iloc[(i - 1) * 49 + (j - 33) * 8 + 1, lc + (j - 29)] for j in range(33,53)]

            K += 1

    return df_rt


# Example usage:
if __name__ == "__main__":
    DInt = 200
    EDInt = 200
    ContInt = 0.01
    TimInt = 5
    Pix = 0.325
    # path = "/mnt/Indapta/clients/211201/eddie_full_actual_625/20231109_MF_Indapta_D2_Dara_PEG_13ET_LP1-01/"
    # path = "/mnt/b/b/InstrumentTestsResults/Results/20240704_Nalm6_NucCART_PDMS_HW-02/"
    # path = "/cellchorus/data/a/InstrumentTests/20240706/Results/eddie/20240706_Nalm6Viability_PDMS_24Hr-01/"
    # path = "/home/eddie/Downloads/testplzzz"
    # path = "/cellchorus/data/c/clients/240201/Dec192024/20241219_TGFBNK_NB1463_NC_8Hr-17/"
    # path = "/cellchorus/data/c/clients/240201/Dec192024/Results/20241219_TGFBNK_NB1463_NC_8Hr-17/"
    path = "/cellchorus/data/b/20250319/20250319_Nalm6_NucCART_Donor2_edge3well-01/"
    labels = ["Eff_x", "Eff_y", "Eff_AR", "Eff_Speed", "Eff_death_int", "Eff_death", "Tar1x", "Tar1y", "Target1_AR",
              "Tar1_Speed", "Tar1D_Int", "Tar1D", "Cont1_Int", "Cont1", "Tar2x", "Tar2y", "Target2_AR", "Tar2_Speed", "Tar2D_Int",
              "Tar2D", "Cont2_Int", "Cont2", "Tar3x", "Tar3y", "Target3_AR", "Tar3_Speed", "Tar3D_Int", "Tar3D", "Cont3_Int", "Cont3",
              "Tar4x", "Tar4y", "Target4_AR", "Tar4_Speed", "Tar4D_Int", "Tar4D", "Cont4_Int", "Cont4", "Tar5x", "Tar5y", "Target5_AR",
              "Tar5_Speed", "Tar5D_Int", "Tar5D", "Cont5_Int", "Cont5", "BEAD_I", "BEAD_II", 'Time_Point']
    result_fields = [
        "UID", "No of targets",
        "tSeek1", "tDeath1", "tContact1", "Cum contact 1", "Absol tar death 1",
        "tSeek2", "tDeath2", "tContact2", "Cum contact 2", "Absol tar death 2",
        "tSeek3", "tDeath3", "tContact3", "Cum contact 3", "Absol tar death 3",
        "tSeek4", "tDeath4", "tContact4", "Cum contact 4", "Absol tar death 4",
        "tSeek5", "tDeath5", "tContact5", "Cum contact 5", "Absol tar death 5",
        "Eff dWell No Cont (um)", "Eff dWell Cont (um)", "Eff AR No Cont", "Eff AR Cont", "tEffDeath",
        "Tar1 dWell No Cont (um)", "Tar1 dWell Cont (um)", "Tar1 AR No Cont", "Tar1 AR Cont",
        "Tar2 dWell No Cont (um)", "Tar2 dWell Cont (um)", "Tar2 AR No Cont", "Tar2 AR Cont",
        "Tar3 dWell No Cont (um)", "Tar3 dWell Cont (um)", "Tar3 AR No Cont", "Tar3 AR Cont",
        "Tar4 dWell No Cont (um)", "Tar4 dWell Cont (um)", "Tar4 AR No Cont", "Tar4 AR Cont",
        "Tar5 dWell No Cont (um)", "Tar5 dWell Cont (um)", "Tar5 AR No Cont", "Tar5 AR Cont"
    ]
    sec = TIMING_secondary(
        path=path, result_fields=result_fields, labels=labels)
    sec.secondary(DInt, EDInt, ContInt, TimInt, Pix)
